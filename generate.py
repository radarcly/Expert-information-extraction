from info.basicInfo import BasicInfo
from info.expertInfo import ExpertInfo
from info.utils import str_is_english,str_is_number
from info.utils import enORGBoundaryRevise,cnORGBoundaryRevise
from info.utils import getORGParticiple
from info.utils import enBoundaryCompletion
from info.utils import reviseIntroducation,cnBoundaryCompletion
from info.utils import enORGIllegalRevise
from info.utils import introducation_is_english_or_chinese
from info.utils import readInputDataByXlrd,readInputDataByOpenpyxl
from info.utils import readPredictResult
from info.utils import initTags
from info.utils import mergeAdjacentTag
from info.utils import basic_edu_work_credit_alignment

import json
import argparse
import xlrd
import openpyxl
import re
import time
import os
import stanza
import copy
import spacy
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

from tqdm import tqdm

# 获取所有tag
def getTags():
    basic_cols = ['name','org','year','month','day','nationality','interest','major','title','post',
                  'depart','sex','nation','nativePlace','address','code','phone','telephone','fax','mail','website']
    work_cols = ['org','title','post','startTime','endTime',
                 'technology','content','depart']
    credit_cols = ['credit','year','level','org','place','category','endYear','grade','reason','content','type']
    education_cols = ['org','startYear','endYear','department','major','education','degree']
    tags = {}
    tags['basic'] = {}
    for basic_col in basic_cols:
        tags['basic'][basic_col] = ["B-"+basic_col.upper(),"I-"+basic_col.upper(),"E-"+basic_col.upper()]
    tags['work'] = {}
    for work_col in work_cols:
        tags['work'][work_col] = ["B-WORK-" + work_col.upper(), "I-WORK-" + work_col.upper(), "E-WORK-" + work_col.upper()]
    tags['credit'] = {}
    for credit_col in credit_cols:
        tags['credit'][credit_col] = ["B-CRE-" + credit_col.upper(),"I-CRE-" + credit_col.upper(),"E-CRE-" + credit_col.upper()]
    tags['edu'] = {}
    for education_col in education_cols:
        tags['edu'][education_col] = ["B-EDU-" + education_col.upper(),"I-EDU-" + education_col.upper(),"E-EDU-" + education_col.upper()]

    return tags

# 初始化输出表格
def initSheets(outwb):
    outwss = []
    basic_cols = ["ID","姓名(中)","姓名(英)","所属机构(中)","所属机构(英)","出生日期","所在国/国籍","研究方向",
    "学科领域","职称","职务","部门/院系","性别","民族","籍贯","地址","邮编","固定电话",
    "手机","传真","邮箱","个人主页","个人介绍"]    
    outws0 = outwb.create_sheet("0基本信息",0)  
    for i in range(len(basic_cols)):
        outws0.cell(1, i+1).value = basic_cols[i] 
    outwss.append(outws0)

    work_cols = ["ID","ord顺序号","机构中","机构英","职称","职务","入职时间","离职时间","技术领域","工作内容","部门"]
    outws1 = outwb.create_sheet("1工作经历",1)  
    for i in range(len(work_cols)):
        outws1.cell(1, i+1).value = work_cols[i] 
    outwss.append(outws1)

    edu_cols = ["ID","ord顺序号","毕业院校(中)","毕业院校(英)","入学年份","毕业年份","院系","专业","学位","学历"]
    outws2 = outwb.create_sheet("2教育经历",2)  
    for i in range(len(edu_cols)):
        outws2.cell(1, i+1).value = edu_cols[i] 
    outwss.append(outws2)

    # cre_cols = ["ID","ord顺序号","荣誉奖项名称","获奖年份","奖励级别","授予机构","授予地", "奖励等级","奖励类别","结束时间",
    # "奖励原因","奖励内容","类型"]
    cre_cols = ["ID","ord顺序号","荣誉奖项名称","获奖年份","奖励级别","授予机构","授予地", "奖励等级","类型","获奖类别"]
    outws3 = outwb.create_sheet("3荣誉奖励",3)  
    for i in range(len(cre_cols)):
        outws3.cell(1, i+1).value = cre_cols[i] 
    outwss.append(outws3)

    return outwss

# 输出基本信息
def writeBasicInfo(basicInfo,outws,row,expert,writeIntro):
    outws.cell(row, 1).value = basicInfo.id
    outws.cell(row, 2).value = basicInfo.cname    
    outws.cell(row, 3).value = basicInfo.ename
    outws.cell(row, 4).value = basicInfo.corganization        
    outws.cell(row, 5).value = basicInfo.eorganization  

    basicInfoDict = dict(expertInfo.basicInfo)  
    # 填基本信息
    basic_info = [expertInfo.basicInfo.getBirthDate(),basicInfoDict["nationality"],basicInfoDict["interest"],
        basicInfoDict["major"], basicInfoDict["title"],basicInfoDict["post"], basicInfoDict["depart"], basicInfoDict["sex"],
        basicInfoDict["nation"],basicInfoDict["nativePlace"],basicInfoDict["address"], basicInfoDict["code"],basicInfoDict["phone"],
        basicInfoDict["telephone"],basicInfoDict["fax"],basicInfoDict["mail"],basicInfoDict["website"],basicInfoDict["introduction"]]
    
    if writeIntro:
        for j in range(6,24):
            #有些非法字符会导致openpyxl写出错
            basic_info[j-6] = ILLEGAL_CHARACTERS_RE.sub(r'', basic_info[j-6])
            outws.cell(row, j).value = basic_info[j-6]
    else:
        for j in range(6,23):
            #有些非法字符会导致openpyxl写出错
            basic_info[j-6] = ILLEGAL_CHARACTERS_RE.sub(r'', basic_info[j-6])
            outws.cell(row, j).value = basic_info[j-6]

# 输出工作信息
def writeWorkInfo(workInfos,outws,work_row_num,id):
    for i in range(len(workInfos)):
        workInfo = workInfos[i]
        outws.cell(work_row_num, 1).value = id
        outws.cell(work_row_num, 2).value = i + 1           
        outws.cell(work_row_num, 3).value = workInfo.cn_org   
        outws.cell(work_row_num, 4).value = workInfo.en_org         
        outws.cell(work_row_num, 5).value = workInfo.title 
        outws.cell(work_row_num, 6).value = workInfo.post 
        outws.cell(work_row_num, 7).value = workInfo.starttime
        outws.cell(work_row_num, 8).value = workInfo.endtime
        outws.cell(work_row_num, 9).value = workInfo.technology
        outws.cell(work_row_num, 10).value = workInfo.content     
        outws.cell(work_row_num, 11).value = workInfo.depart
        work_row_num += 1
    return work_row_num

# 输出教育信息
def writeEducationInfo(educationInfos,outws,edu_row_num,id):
    for i in range(len(educationInfos)):
        educationInfo = educationInfos[i]       
        outws.cell(edu_row_num, 1).value = id 
        outws.cell(edu_row_num, 2).value = i + 1 
        outwss[2].cell(edu_row_num, 3).value = educationInfo.cn_org 
        outwss[2].cell(edu_row_num, 4).value = educationInfo.en_org 
        outws.cell(edu_row_num, 5).value = educationInfo.startyear       
        outws.cell(edu_row_num, 6).value = educationInfo.endyear        
        outws.cell(edu_row_num, 7).value = educationInfo.department              
        outws.cell(edu_row_num, 8).value = educationInfo.major        
        outws.cell(edu_row_num, 9).value = educationInfo.degree  
        outws.cell(edu_row_num, 10).value = educationInfo.education
        edu_row_num += 1 
    return edu_row_num

# 输出荣誉获奖
def writeCreditInfo(creditInfos,outws,credit_row_num,id):
    for i in range(len(creditInfos)):
        creditInfo = creditInfos[i]
        outws.cell(credit_row_num, 1).value = id  
        outws.cell(credit_row_num, 2).value = i + 1              
        outws.cell(credit_row_num, 3).value = creditInfo.credit 
        outws.cell(credit_row_num, 4).value = creditInfo.year 
        outws.cell(credit_row_num, 5).value = creditInfo.level
        outws.cell(credit_row_num, 6).value = creditInfo.org
        outws.cell(credit_row_num, 7).value = creditInfo.place
        outws.cell(credit_row_num, 8).value = creditInfo.grade
        outws.cell(credit_row_num, 9).value = creditInfo.type
        outws.cell(credit_row_num, 10).value = creditInfo.category
        credit_row_num += 1
    return credit_row_num

parser = argparse.ArgumentParser()
parser.add_argument("--expertData",
                    default="data/data.xlsx",
                    type=str)
parser.add_argument("--predictResult",
                    default="predict_result/predict_result_cn.json",
                    type=str)  
parser.add_argument("--nerResult",
                    default="ner_result/stanza_cn_ner.json",
                    type=str)  
parser.add_argument("--isRevised",
                    default=True,                    
                    type=bool)
parser.add_argument("--isDebug",
                    default=False,                    
                    type=bool)
parser.add_argument("--writeWork",
                    default=True,                    
                    type=bool)
parser.add_argument("--writeEdu",
                    default=True,                    
                    type=bool)
parser.add_argument("--writeCre",
                    default=True,                    
                    type=bool)
parser.add_argument("--writeIntro",
                    default=True,                    
                    type=bool)


# parser.add_argument("--writeWork",
#                     default=False,                    
#                     type=bool)
# parser.add_argument("--writeEdu",
#                     default=False,                    
#                     type=bool)
# parser.add_argument("--writeCre",
#                     default=False,                    
#                     type=bool)
# parser.add_argument("--writeIntro",
#                     default=False,                    
#                     type=bool)



# 相邻合并超参                    
parser.add_argument("--mergeCNTagDistance",
                    default=3,
                    type=int)
# 判断是同一条记录超参                
parser.add_argument("--cnTheSameRecordDistance",
                    default=10,
                    type=int)    
args = parser.parse_args()

tags = getTags()
print("正在读取所有专家信息")
experts = readInputDataByXlrd(args.expertData)
print("读取完所有专家信息,正在读取模型预测结果")
results = readPredictResult(args.predictResult)
print("读取完模型预测结果,正在读取stanza结果")                  
ner_results = getORGParticiple(args.nerResult)
print("读取完stanza结果，正在解码")    

# en_chunking_nlp = spacy.load("en")
outwb = openpyxl.Workbook()  # 打开一个将写的文件

# 初始化所有表格第一行
outwss = initSheets(outwb)

# 初始化三个全局变量 分别存work的行数，credit的行数和edu的行数
work_row_num = credit_row_num = edu_row_num = same_row_num = row = 2


with open("text1.txt", "w", encoding='utf-8') as f:
    # 后台运行不需要进度条
    for key,value in results.items():
    # 前台运行需要进度条
    # for key,value in tqdm(results.items()):
        if key != "":
        # if key in ["10046237000"]:
            language,introducation = introducation_is_english_or_chinese(experts[key]['简介'])
            tags = initTags(key,introducation,value[0],value[1])
            # for j in range(len(value[0])):
            #     print(str(j) + "\t" + value[0][j] +"\t" + value[1][j])
            # for j in range(len(tags)):
            #     print(str(j) + "\t" + introducation[j] +"\t" + tags[j])
            origin_tags = copy.deepcopy(tags)
            if len(tags) > 0:           
                # 进行中文分词边界调整
                tags = cnBoundaryCompletion(introducation,tags,args.isDebug)
                # 进行中文机构NER边界调整
                tags = cnORGBoundaryRevise(introducation,tags,ner_results[key],args.isDebug)
                # 循环遍历一次tag 将相邻tag进行合并
                tags = mergeAdjacentTag(introducation,tags,args.mergeCNTagDistance,args.isDebug)
                # print(key,"完成合并")
                # 循环遍历一次tag 解决一些错位的情况 如工作机构预测为教育机构
                tags = basic_edu_work_credit_alignment(introducation,tags,args.isDebug)
                tags = basic_edu_work_credit_alignment(introducation,tags,args.isDebug)

            expertInfo = ExpertInfo(key,language,experts[key]['简介'],introducation,args.cnTheSameRecordDistance)   
            expertInfo.decoder(list(introducation),tags)   
            if args.isRevised:
                expertInfo.revise(experts[key]) 


        writeBasicInfo(expertInfo.basicInfo,outwss[0],row,experts[key],writeIntro = args.writeIntro)

        if (args.writeWork) :
            # 工作经历
            work_row_num = writeWorkInfo(expertInfo.workInfos,outwss[1],work_row_num,key)
        
        if (args.writeEdu) :
            # 教育经历
            edu_row_num = writeEducationInfo(expertInfo.educationInfos,outwss[2],edu_row_num,key)
    
        if (args.writeCre) :
            # 最后荣誉获奖
            credit_row_num = writeCreditInfo(expertInfo.creditInfos,outwss[3],credit_row_num,key)
        
        # 行数+1 下一个学者
        row = row + 1  

# 保存预测文件    
str = time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
print("正在保存结果文件")   
outwb.save('prediction_result_' + str + '.xlsx')
print("保存完毕")   

    
    

