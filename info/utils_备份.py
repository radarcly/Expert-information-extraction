import jieba
import xlrd
import openpyxl
import re
import string
import json
from zhon.hanzi import punctuation
from tqdm import tqdm
# 判断一个字符是不是英文字符
def char_is_alphabet(uchar):
    """判断一个unicode是否是英文字母"""
    if (u'\u0041' <= uchar <= u'\u005a') or (u'\u0061' <= uchar <= u'\u007a'):
        return True
    else:
        return False

# 判断一个字符是不是数字
def char_is_number(uchar):
    """判断一个unicode是否是数字"""
    if u'\u0030' <= uchar <= u'\u0039':
        return True
    else:
        return False

# 判断一个字符是不是汉字
def char_is_chinese(uchar):
    """判断一个unicode是否是汉字"""
    if  u'\u4e00' <= uchar <= u'\u9f5a':
        return True
    else:
        return False

# 判断一个字符是不是括号
def char_is_prentheses(uchar):
    if uchar in ["(",")","（","）"]:
        return True
    else:
        return False

# 判断一个字符是不是单引号或者双引号
def char_is_quotations(uchar):
    if uchar in ["”","“","‘","’","'","\""]:
        return True
    else:
        return False

# 判断一个字符是不是空格
def char_is_space(uchar):
    """判断一个unicode是否是空字符串（包括空格，回车，tb）"""
    space = [u'\u0020', u'\u000A', u'\u000D', u'\u0009']
    if uchar in space:
        return True
    else:
        return False

# 判断一个字符是不是标点符号
def char_is_punctuation(uchar):
    if uchar in string.punctuation or uchar in punctuation:
        return True
    else:
        return False

# 判断一个字符是否非汉字，数字，空字符和英文字符还不是标点符号
def char_is_other(uchar):
    """判断是否非汉字，数字，空字符和英文字符"""
    if not (char_is_space(uchar) or char_is_chinese(uchar) or char_is_number(uchar) or char_is_alphabet(uchar) or char_is_punctuation(uchar) ):
        return True
    else:
        return False

# 判断一个字符串是不是英文字符串，如果一个字符串英文字符长度是中文字符长度5倍就认为他是英文字符串，并且英文字符长度不为0
def str_is_english(str):
    e_num = 0
    c_num =0
    for char in str:
        if char_is_alphabet(char):
            e_num += 1
        elif char_is_chinese(char):
            c_num += 1
    if e_num > 5*c_num:
        return True
    else:
        return False

# 判断一个字符串是不是中文字符串，如果一个字符串英文字符长度小于中文字符长度5倍就认为他是中文字符串，并且中文字符长度不为0
def str_is_chinese(str):
    e_num = 0
    c_num =0
    for char in str:
        if char_is_alphabet(char):
            e_num += 1
        elif char_is_chinese(char):
            c_num += 1
    if  5 *c_num >= e_num:
        return True
    else:
        return False

# 判断一个字符串是否不含英文字母
def str_without_alphabet(str):
    for char in str:
        if char_is_alphabet(char):
            return False
    return True

# 判断一个字符串是否是int
def str_is_number(str):
    try:
        int(str)
        return True
    except ValueError:
        pass

# 得到一个字符串中字母的数量
def getStringlphbetNumber(str):
    num = 0
    for char in str:
        if char_is_alphabet(char):
            num += 1
    return num

# 得到一个字符串中汉字的数量
def getStringChineseNumber(str):
    num = 0
    for char in str:
        if char_is_chinese(char):
            num += 1
    return num

# 判断专家介绍为英文或中文，同时将长介绍修正
def introducation_is_english_or_chinese(introducation):
    lnguge = ""
    if str_is_chinese(introducation[:100]) or getStringChineseNumber(introducation) > 50 :
        lnguge = "chinese"
    else:
        lnguge = "english"
    if len(introducation) > 4000:
        introducation = introducation[0:4000]
    return lnguge,introducation

# 通过xlrd工具读取所有专家数据，适用于xls后缀文件,第二个参数为想要读取专家数量，默认为全部
def readInputDataByXlrd(filename,experts_num=-1):
    xlsx = xlrd.open_workbook(filename)
    sheet = xlsx.sheets()[0]
    rows = sheet.nrows
    cols = sheet.ncols

    # 如果没有指定专家数量,则返回所有专家
    if experts_num == -1:        
        experts_num = rows   

    key_list = []
    # 获得key_list id 名字 所在机构 简介
    for i in range(cols):
        key_list.append(sheet.cell(0, i).value)
    # 获得experts
    experts = {}
    for i in tqdm(range(1, experts_num)):
        expert_id =  str(sheet.cell(i, 0).value)
        expert = {}
        # print(sheet.cell(i, 0).value)
        introducation = ""
        # # 获得专家 id 名字 所在机构信息
        
        for j in range(0, 3):
            expert[key_list[j]] = str(sheet.cell(i, j).value)

        # # 专家的简介信息需要合并第四列到第九列
        for j in range(3, cols):
            if type(sheet.cell(i, j).value) == float:
                introducation += str(int(sheet.cell(i, j).value))
            else:
                introducation += str(sheet.cell(i, j).value)
        # 用正则表达式删除html标签，之后删掉&nbsp,换行符和制表符
        expert[key_list[3]] = reviseIntroducation(introducation)
        expert["原始简介"] = introducation
        experts[expert_id] = expert
    return experts

# 通过openpyxl工具读取所有专家数据，适用于xlsx后缀文件，第二个参数为想要读取专家数量，默认为全部
def readInputDataByOpenpyxl(filename,experts_num=-1):
    # 读取excel
    workbook = openpyxl.load_workbook(filename)
    # 获取所有sheet的名字
    sheetnmes = workbook.sheetnmes 
    # 加载第一个sheet
    sheet = workbook[sheetnmes[0]] 
    # 获取sheet的行数和列数
    sheet_rows = sheet.mx_row
    sheet_cols = sheet.mx_column

    # 如果没有指定专家数量,则返回所有专家
    if experts_num == -1:        
        experts_num = sheet_rows
    
    # 获得key_list id 名字 所在机构 简介 openpyxl下标从1开始
    key_list = []  
    for i in range(1,sheet_cols+1):
        key_list.append(sheet.cell(1, i).value)
    
    # 获得experts
    experts = {}
    for i in tqdm(range(2, experts_num + 1)):
        expert_id =  str(sheet.cell(i, 1).value)
        expert = {}
        # print(sheet.cell(i, 0).value)
        introducation = ""
        # # 获得专家 id 名字 所在机构信息        
        for j in range(1, 4):
            expert[key_list[j]] = str(sheet.cell(i, j).value)
        # # 专家的简介信息需要合并第四列到第九列
        for j in range(4,sheet_cols + 1):
            if type(sheet.cell(i, j).value) == float:
                introducation += str(int(sheet.cell(i, j).value))
            elif str(sheet.cell(i, j).value) != "None":
                introducation += str(sheet.cell(i, j).value)
        # 用正则表达式删除html标签，之后删掉&nbsp,换行符和制表符
        expert[key_list[3]] = reviseIntroducation(introducation)
        expert["原始简介"] = introducation
        experts[expert_id] = expert
    return experts

# 通过json文件读取所有专家数据
def readInputDataByJson(filename):
    with open(filename,'r',encoding='utf8') as f:
        results = json.load(f)
    return results

# 读取预测结果
def readPredictResult(filename):
    with open(filename,'r',encoding='utf8') as f:
        results = json.load(f)
    return results

# 读取中文ner结果
def getCNORGParticiple(filename):
    with open(filename,'r',encoding='utf8') as f:
        results = json.load(f)
    return results

# 测试数据
def getTestData():
    experts = {}
    expert = {}
    expert['id'] = "1"
    expert['简介'] = "李冉，男，复旦大学教授，1979年10月1日生，国籍：中国，汉族，江苏南京人，担任马克思主义学院院长；马克思主义研究院副院长，" \
                             "研究方向：马克思主义，学科领域，政治学，通讯地址光华楼101，邮编：201101，电话456789，手机123456" \
                             "传真88880，邮箱88888@qq.com，个人网页www.bidu.com" \
                   "教育经历：南京大学本科研究生连读 1995年进入清华大学机械工程系修读液压传动及气动博士 2000年毕业" \
                   "获奖经历 2010，2011连续两年获得长江学者称号，2012年获得世界级别的由哈佛大学在美国颁发的持续到2014年的教学成果奖金奖" \
                   "工作经历 2016年在北京大学国际关系学院任院长，同时在物理学院任副院长"
    experts[expert['id']] = expert

    # mrked_experts = {}
    # mrked_expert = {}
    # mrked_expert['cnme'] = "李冉"
    # mrked_expert['enme'] = ""
    # mrked_expert['corgniztion'] = "复旦大学"
    # mrked_expert['eorgniztion'] = ""
    # mrked_expert['title'] = "教授"
    # mrked_expert['year'] = "1979"
    # mrked_expert['month'] = "10"
    # mrked_expert['dy'] = "1"
    # mrked_expert['sex'] = "男"
    # mrked_expert['ntionlity'] = "中国"
    # mrked_expert['ntion'] = "汉族"
    # mrked_expert['ntivePlce'] = "江苏南京"
    # mrked_expert['deprt'] = "马克思主义学院；马克思主义研究院"
    # mrked_expert['post'] = "马克思主义学院院长；马克思主义研究院副院长"
    # mrked_expert["interest"] = "马克思主义"
    # mrked_expert["mjor"] = "政治学"
    # mrked_expert["ddress"] = "光华楼101"
    # mrked_expert["code"] = "201101"
    # mrked_expert["phone"] = "456789"
    # mrked_expert["telephone"] = "123456"
    # mrked_expert["fx"] = "88880"
    # mrked_expert["mil"] = "88888@qq.com"
    # mrked_expert["website"] = "www.bidu.com"

    # mrked_expert["eduction"] = []
    # mrked_expert["eduction"].append({"cschool": "南京大学", "eduction": "本科"})
    # mrked_expert["eduction"].append({"cschool": "南京大学", "eduction": "研究生"})
    # mrked_expert["eduction"].append({"cschool": "清华大学", "strtYer": "1995","endYer":"2000","deprtment":"机械工程系",
    #                                     "mjor":"液压传动及气动","degree":"博士"})
    # mrked_expert["credit"] = []
    # mrked_expert["credit"].append({"credit": "长江学者", "year": "2010"})
    # mrked_expert["credit"].append({"credit": "长江学者", "year": "2011"})
    # mrked_expert["credit"].append({"credit": "教学成果奖", "year": "2012","level":"世界级","orgniztion":"哈佛大学",
    #                                 "plce":"美国","endYer":"2014","grde":"金奖"})
    # mrked_expert["work"] = []
    # mrked_expert["work"].append({"corgniztion": "北京大学", "deprt":"国际关系学院","strtTime": "2016","post":"国际关系学院院长"})
    # mrked_expert["work"].append({"corgniztion": "北京大学", "deprt":"物理学院","strtTime": "2016","post":"物理学院副院长"})
    # mrked_expert["credit"].append({"credit": "教学成果奖", "year": "2012", "level": "世界级", "orgniztion": "哈佛大学",
    #                                 "plce": "美国", "endYer": "2014", "grde": "金奖"})
    # mrked_experts[expert['id']] = mrked_expert
    return experts

def delete_UNK_And_Symbol(str):
    str = str.strip()
    str = str.replace("[UNK]","")
    return str

def check_len_more_than_one(str):
    if len(str) < 2:
        str = ""
    return str

def reviseTime(str_time):
    # 如果字符串中有年和月
    if "年" in str_time and "月" in str_time:
        index1 = str_time.find("年")
        index2 = str_time.find("月")
        # 找到年和月
        year = str_time[0:index1]
        month = str_time[index1+1:index2]
        # 如果月份以0打头删掉
        if(month.startswith("0")):
            month = month[1:]
        # 年份是数字，月份是数字，且年份大于1800小于2025 月份 大于0小于13 才返回
        if str_is_number(year) and str_is_number(month) and int(year) > 1800 and int(year) < 2025 and int(month) > 0 and int(month) < 13:
                return year + "年" + month + "月"
        else:
            return ""
    # 如果字符串只有年
    elif "年" in str_time:
        index1 = str_time.find("年")
        # 找到年
        year = str_time[0:index1]
        # 如果年后面还有东西则代表可能有月信息，提取月信息
        if len(year) == 4 and len(str_time) > index1  + 1 :
            month = str_time[index1+1]
            # 如果月信息和年信息格式都正确
            if str_is_number(year) and str_is_number(month) and int(year) > 1800 and int(year) < 2050 and int(month) > 0 and  int(month) < 13:
                return year + "年" + month + "月"
            # 否则年信息格式正确，只输出年
            elif str_is_number(year) and int(year) > 1800 and int(year) < 2050:
                return year + "年"
            else:
                return ""
        # 否则只有年信息，保证年份为1800-2050之间的数字即可
        elif str_is_number(year) and int(year) > 1800 and int(year) < 2050 :
            return year + "年"
        else:
            return ""
    # 如果字符串是单纯的四位数字
    elif str_is_number(str_time) and int(str_time) > 1800 and int(str_time) < 2050 :
        return str_time + "年"
    # 如果字符串以点分隔年月
    elif "." in str_time:
        arr = str_time.split(".")
        if(arr[1].startswith("0")):
            arr[1] = arr[1][1:]
        year = arr[0]
        month =arr[1]
         # 如果月信息和年信息格式都正确
        if str_is_number(year) and str_is_number(month) and int(year) > 1800 and int(year) < 2050 and int(month) > 0 and int(month) < 13:
                return year + "年" + month + "月"
        # 否则年信息格式正确，只输出年
        elif str_is_number(year) and int(year) > 1800 and int(year) < 2050:
                return year + "年"
        else:
            return ""
    else:
        return ""

def checkTime(strtTime,endTime):

    # 如果有一个时间为空直接返回
    if strtTime == "" or endTime == "":
        return endTime
    
    # 找到年份，
    strtYer = int(strtTime[0:strtTime.find("年")])
    endYer = int(endTime[0:endTime.find("年")])
    if endYer > strtYer:
        return endTime
    elif endYer < strtYer:
        return ""
    
    # 年份相同的情况下，如果都有月份
    if "月" in strtTime and "月" in endTime:
        strtMonth = int(strtTime[strtTime.find("年") +1 :strtTime.find("月")])
        endMonth = int(endTime[endTime.find("年") +1 :endTime.find("月")])
        if endMonth > strtMonth:
            return endTime
        else:
            return ""
    # 如果都没月份，则判断成立
    if "月" not in strtTime and "月" not in endTime:
        return endTime
    
    # 如果一个有月份一个没月份，默认出错，endTime清零
    return ""
   
def get_element(tokens,tags,i,tg):
    strt = i
    end = i + 1
    res = tokens[strt]
    while end < len(tokens) and tags[end] == tg:
        res += tokens[end]
        end = end + 1
    return res,tg,end

def delete_UNK(intro_list):
    for intro in intro_list:
        intro = intro.replace("[UNK]","")

# 相邻合并
def mergeAdjacentTag(introducation,tags,distnce,debug=False):
    lstIndex = 0
    lstTg = tags[0]
    for i in range(len(tags)):
        # 中间只有中文,英文字母，数字,空格，或者括号，或者-才合并
        if tags[i] != "" and tags[i] == lstTg and i - lstIndex <= distnce and i - lstIndex > 1 and checkMergeCondition(lstTg,introducation,tags,lstIndex+1,i):
            for j in range(lstIndex+1,i):
                if debug:
                    print("相邻合并",j)
                tags[j] = lstTg                  
            lstIndex = i
            lstTg = tags[i]
        elif tags[i] != "":
            lstIndex = i
            lstTg = tags[i]
    return tags

# 检查是否满足合并条件
def checkMergeCondition(tg,introducation,tags,begin,end):
    for i in range(begin,end):        
        char = introducation[i]
        if char_is_chinese(char) :
            continue
        elif char_is_alphabet(char):        
            continue
        elif char_is_number(char):
            continue
        elif char_is_space(char):
            continue
        elif char_is_prentheses(char):
            continue
        elif char_is_quotations(char):
            continue
        # 如果要补充的tg为NME并且chr为.的话
        elif tg == "NME" and char == ".":
            continue
        elif char == "—":
            continue
        else:
            return False
    return True

# 修复中英文括号，引号，书名号不对齐相关问题
def reviseSymbolPair(s):
    # print(s)
    # 删掉最后多余括号
    if len(s) > 0 and s[0] in [")","）"]:
        s = s[1:]
    if len(s) > 0 and s[-1] in ["(","（"]:
        s = s[:-1]
        
    # 补充括号
    if s.count("(") - s.count(")") == 1:
        s = s +")"
    elif s.count("（") - s.count("）") == 1:
        s = s +"）"
    elif s.count("(") - s.count(")") == -1:
        s = "(" + s 
    elif s.count("（") - s.count("）") == -1:
        s = "（" + s 
    
    # 补充引号
    if s.count("“") - s.count("”")  == 1:
        s = s + "”"
    elif s.count("“") - s.count("”") == -1:
        s = "“" + s
    
    # 补充书名号
    if s.count("《") - s.count("》")  == 1:
        s = s + "》"
    elif s.count("《") - s.count("》") == -1:
        s = "《" + s

    return s 

# 修复一些乱七八糟的前缀和后缀
def revisePrefixAndSuffix(s):
    delete_prefix_dict =    [",","，", 
                              ".","。",
                              ":","：",
                              "、","/",
                              "-","——",
                              "&","'",
                              "及","与","获","于","和","年","的","任"]
    delete_suffix_dict =    [",","，", 
                              ".","。",
                              ":","：",
                              "、","/",
                              "-","——",
                              "&","'",
                              "及","与","获","于","和","年","的","任"]
        
    # 删除指定前缀
    for prefix in delete_prefix_dict:
        if s.startswith(prefix):
            s= s[len(prefix):]
            break
        
    # 删除指定后缀
    for suffix in delete_suffix_dict:
        if s.endswith(suffix):
            s = s[:-len(suffix)]
            break

    return s.strip()



# 检查从begin 到 end 中的token有没有出现句号
def checkFullStopInTokens(tokens,begin,end):
    for i in range(begin,end):
        if tokens[i] in [".","。"]:
            return True
    return False

def reviseEnglishItem(s):
    s = s.strip()
    s_copy = s
    s = s.lower()
    stop_words = ['i', 'me', 'my', 'myself', 'we', 'our', 'ours', 'ourselves', 'you', 'your', 'yours', 'yourself', 'yourselves', 
                  'he', 'him', 'his', 'himself', 'she', 'her', 'hers', 'herself', 'it', 'its', 'itself', 'they', 'them', 'their', 
                  'theirs', 'themselves', 'wht', 'which', 'who', 'whom', 'this', 'tht', 'these', 'those', 'm', 'is', 're', 'ws', 
                  'were', 'be', 'been', 'being', 'hve', 'hs', 'hd', 'hving', 'do', 'does', 'did', 'doing', '', 'n', 'the', 'and', 
                  'but', 'if', 'or', 'becuse', 's', 'until', 'while', 'of', 't', 'by', 'for', 'with', 'bout', 'ginst', 'between', 
                  'into', 'through', 'during', 'before', 'fter', 'bove', 'below', 'to', 'from', 'up', 'down', 'in', 'out', 'on',
                  'off', 'over', 'under', 'gin', 'further', 'then', 'once', 'here', 'there', 'when', 'where', 'why', 'how', 'll', 
                  'ny', 'both', 'ech', 'few', 'more', 'most', 'other', 'some', 'such', 'no', 'nor', 'not', 'only', 'own', 'sme',
                  'so', 'thn', 'too', 'very', 's', 't', 'cn', 'will', 'just', 'don', 'should', 'now', 'd', 'll', 'm', 'o', 're', 
                  've', 'y', 'in', 'ren', 'couldn', 'didn', 'doesn', 'hdn', 'hsn', 'hven', 'isn', 'm', 'mightn', 'mustn', 'needn',
                  'shn', 'shouldn', 'wsn', 'weren', 'won', 'wouldn','a','an']
    for word in stop_words:
        if s.startswith(word):
            s_copy = s_copy[len(word):]
        if s.endswith(word):
            s_copy = s_copy[:-len(word)]
    s_copy = s_copy.strip()
    if len(s_copy .split()) < 2:
        s_copy = ""
    return s_copy


# def mergeENdjcentTg(introducation,tags,distnce):
#     lstIndex = 0
#     lstTg = tags[0]
#     for i in range(len(tags)):
#         if tags[i] != "" and tags[i] == lstTg and i - lstIndex <= distnce and i - lstIndex > 1:
#             for j in range(lstIndex+1,i):
#                 # 是中文或者英文字符才合并
#                 if char_is_chinese(introducation[j]) or char_is_alphabet(introducation[j]) or char_is_space(introducation[j]):
#                     # print("相邻合并",j)
#                     tags[j] = lstTg                  
#             lstIndex = i
#             lstTg = tags[i]
#         elif tags[i] != "":
#             lstIndex = i
#             lstTg = tags[i]
#     return tags

def cnBoundaryCompletion(introducation,tags,debug):
    # doc = zh_nlp(introducation[0:100])
    begin_indexz,end_indexz = getCNPrticiple(introducation)
    # print(begin_indexz)
    # print(end_indexz)
    # 得到所有左边界和右边界
    # for sentence in doc.sentences:
    #     for token in sentence.tokens:
    #         # print(token)
    #         begin_indexz.append(token.strt_chr)
    #         end_indexz.append(token.end_chr)
    # print(begin_indexz)
    # print(end_indexz)

    # 循环遍历第二个到倒数第二个tg
    for i in range(1,len(tags)-1):
        # 第一种情况如果tg不等空且上一个tg不等于这个tg，代表左边界的情况出现了
        if tags[i] != "" and tags[i-1] != tags[i]:
            # 如果i不在中文分词左边界，找到列表中小于i最大的左边界
            left_boundry = getLeftBoundry(i,begin_indexz)
            # 修改tg
            tags = modifyTag(debug,introducation,tags,left_boundry,i,tags[i],"左边界")                     
        # 如果tg不等于空，且下一个tg不等于这个tg，代表右边界的情况出现了
        if tags[i] != "" and tags[i+1] != tags[i]:
            # 如果i+1不在中文分词右边界，找到列表中大于i+1最小的右边界      
            right_boundry = getRightBoundry(i+1,end_indexz)
            tags = modifyTag(debug,introducation,tags,i+1,right_boundry,tags[i],"右边界")

    return tags

def enBoundaryCompletion(introducation,tags,en_chunking_nlp):
    # doc = zh_nlp(introducation[0:100])
    chunking_index_list = getENPrticiple(introducation,en_chunking_nlp)
    # print(chunking_index_list)
    for i in range(1,len(tags)-1):
        if judgeWhetherInORGPticiple(chunking_index_list,i) != -1:           
            left_boundry,right_boundry  =chunking_index_list[judgeWhetherInORGPticiple(chunking_index_list,i)]
            modifyTag(False,introducation,tags,left_boundry,right_boundry,tags[i],"ner修改",["TITLE","POST","MJOR","DEPRT","WORK-DEPRT","EDU-DEPRTMENT"])
            i = right_boundry - 1        
    return tags
   
def cnORGBoundaryRevise(introducation,tags,org_index_list,debug):
    for i in range(1,len(tags)-1):
        # 先做stnz ner 边界补全
        if tags[i] in ["ORG","WORK-ORG","EDU-ORG"] and judgeWhetherInORGPticiple(org_index_list,i) != -1:           
            left_boundry,right_boundry = org_index_list[judgeWhetherInORGPticiple(org_index_list,i)]
            modifyTag(debug,introducation,tags,left_boundry,right_boundry,tags[i],"ner修改",["TITLE","POST","MJOR","DEPRT","WORK-DEPRT","EDU-DEPRTMENT"])
            i = right_boundry - 1
    return tags

# 检查通过
def enORGBoundaryRevise(introducation,tags,org_index_list):
    i = 1
    while i < len(tags) - 1:
        if tags[i] in ["ORG","WORK-ORG","EDU-ORG"] and judgeWhetherInORGPticiple(org_index_list,i) != -1:           
            left_boundry,right_boundry = org_index_list[judgeWhetherInORGPticiple(org_index_list,i)]
            modifyTag(False,introducation,tags,left_boundry,right_boundry,tags[i],"ner修改",["TITLE","POST","MJOR","DEPRT","WORK-DEPRT","EDU-DEPRTMENT"])
            i = right_boundry  
        else:
            i = i + 1       
    return tags

def modifyTag(debug,introducation,tags,begin,end,tg,hint,tg_list=[]):
    # 先检查要修改的tg有没有包含tg_list中的tg，如果包含直接整个不修改
    flg = False
    for i in range(begin,end):
        if tags[i] in tg_list:
            flg = True
            break
    # print(flg)
    if  not flg:
        for i in range(begin,end):
            tags[i] = tg
            if debug:
                print("modify"+hint,i,tags[i])
    return tags

def getLeftBoundry(index,indexz):
    if index not in indexz:
        # 特殊情况index比indexz最后一个index还大，那说明左边界就是最后一个index
        if index > indexz[-1]:
            return indexz[-1]
        # 其他情况遍历indexz中所有index，如果当前小，下一个大则代表为当前
        for i in range(len(indexz)-1):            
            if indexz[i] < index and indexz[i+1]> index:
                return indexz[i]
    return index

def getRightBoundry(index,indexz):
    if index not in indexz:
        # 特殊情况 index比indexz中第一个index还小，那说明右边界就是第一个index
        if index < indexz[0]:
            return indexz[0]
        for i in range(len(indexz)-1):
            if indexz[i] < index and indexz[i+1]> index:
                return indexz[i+1]
    return index


# 将指定begin到end index的tg清空，修正英文组织和荣誉奖励时候用到
def deleteTag(tags,begin,end,hint):
    for i in range(begin,end):
        tags[i] = ""
    return tags

# 初始化tg 英文中文都用这个方法
def initTags(key,introducation,tokens,origin_tgs):
    index = 0
    tags = []
    i = 0
    while i < len(tokens):
        # [UNK] 对应的并不一定是一位
        if(tokens[i] == "[UNK]"):            
            j = 1
            # 可能存在连续的[UNK],找到后面不是[UNK]的词
            while (i+j) < len(tokens) and tokens[i+j] == "[UNK]":
                j = j + 1
            # 如果找到了不是UNK的token
            if (i+j) < len(tokens):
                #从index 开始找这个token所在的位置
                end_index = introducation.lower().find(tokens[i+j],index)
                for k in range(index,end_index):
                    tags.append("")
                # print(index,end_index,tokens[i+j])
                index = end_index
            # 说明UNK一直连到末尾            
            else:
                for k in range(index,len(introducation)):
                    # tags.append("UNK")
                    tags.append("")
                index = len(introducation)
            i = i + j        
        else:
            if(introducation[index:index+len(tokens[i])].lower() != tokens[i]):
                print(key)
                print(introducation)
                # print(len(tokens))
                print(str(index) +"\t" + str(i))
                print(str(len(tokens[i])) + "\t"  + tokens[i])
                for j in range(len(tokens[i])):
                    print(tokens[i][j])
                print(str(len(introducation[index:index+len(tokens[i])].lower())) + "\t" + introducation[index:index+len(tokens[i])].lower())
                for j in range(len(origin_tgs)):
                    print(str(j) + "\t" +introducation[j] + "\t" + tokens[j] +"\t" + origin_tgs[j])
                
                # print(str(index) +"\t" + str(i) + "\t" + introducation[index:index+len(tokens[i])].lower() + "\t" + tokens[i])
                exit()
            for j in range(len(tokens[i])):           
                tags.append(origin_tgs[i])
            index += len(tokens[i])
            i += 1
       
        # print(index,introducation[index])
        # 一个token结束之后每有一个空格就加一个
        while index < len(introducation) and introducation[index] == " ":
            tags.append("")
            index += 1
    return tags

def getCNPrticiple(introducation):
    begin_indexz = []
    end_indexz = []
    jieba.load_userdict("/home/cly/poc/dict/cn_vocbulry.txt")
    seg_list = jieba.lcut(introducation, cut_ll=False)
    index = 0
    for seg in seg_list:
        begin_indexz.append(index)
        index = index + len(seg)
        end_indexz.append(index)
    return begin_indexz,end_indexz


    
def getENORGPrticiple(introducation,en_nlp):
    org_index_list = []
    doc = en_nlp(introducation)
    for sent in doc.sentences:
        for ent in sent.ents:
            if ent.type == "ORG":
                # print(ent)
                org_index_list.append((ent.strt_chr,ent.end_chr))
    return org_index_list

def judgeWhetherInORGPticiple(org_index_list,index):
    for i in range(len(org_index_list)):
        org_index = org_index_list[i]
        if index >= org_index[0] and index < org_index[1]:
            return i
    return -1

# def chrTokenizerToBertTokenizer(intro_list,tg_list):
#     index = 0
#     tags = []    
#     for intro in intro_list:
#         tags.append(tg_list[index])
#         index += len(intro) 
#     return tags

# 英文chunking的时候是组合名词短语，但注意要同一语系，以空格做分割，判断空格前和空格后是否属于同一语种
# 属于则分割，不属于则分开
def getENPrticiple(introducation,en_chunking_nlp):
    end_index = 0
    chunking_index_list = []
    doc = en_chunking_nlp(introducation)
    for chunk in doc.noun_chunks:
        current_chunk_list = chunk.text.split()
        # print(current_chunk_list)
        revised_chunk_list = []
        lnguge = " "
        for chunk_list_element in current_chunk_list:
            # 第一种情况 再识别第一个元素，判断他是中文还是英文，之后直接将结果添加进文本中
            if lnguge == " " :
                lnguge = "en" if str_is_english(chunk_list_element) else "cn" 
                revised_chunk_list.append(chunk_list_element)
            # 第二种情况，上一个元素为英文，当前元素也为英文，直接将当前元素加到上个元素后面         
            elif lnguge == "en" and str_is_english(chunk_list_element):
                revised_chunk_list[-1] += " " + chunk_list_element
            # 第三种情况，上一个元素为中文，当前元素也为中文，直接将当前元素加到上个元素后面         
            elif lnguge == "cn" and not str_is_english(chunk_list_element):
                # print(chunk_list_element)
                revised_chunk_list[-1] += " " + chunk_list_element
            # 最后一种情况，当前元素和上一个元素不一致的情况
            else:
                lnguge = "en" if lnguge == "cn" else "cn"
                revised_chunk_list.append(chunk_list_element)
        for chunk_element in revised_chunk_list:
            # print(chunk_element)
            begin_index = introducation.find(chunk_element,end_index)
            end_index = begin_index + len(chunk_element)
            chunking_index_list.append((begin_index,end_index))
    return chunking_index_list

#  修正个人介绍
def reviseIntroducation(introducation): 
    # 去掉html标签
    introducation = re.sub('<[^<]+?>', '', introducation)
    # 去掉换行符
    introducation = introducation.replace("\n"," ").replace("\r"," ")
    # 去掉 "\t"
    introducation = introducation.replace("\t"," ")
    # 去掉&nbsp
    introducation = introducation.replace("&nbsp;"," ")
    # 去除多余空格
    introducation = spaceReplace(introducation)
    # 替换一些奇怪语言
    introducation = strangeSymbolReplace(introducation)
    # 删掉两头空白
    introducation = introducation.strip()
    return introducation

# 将字符串中得多个空格替换为一个空格
def spaceReplace(s): 
    return " ".join(s.split())


def strangeSymbolReplace(s):
    introducation = ""
    for uchar in s:
        if not char_is_other(uchar):
            introducation += uchar
        # else:
        #     print(uchar)
    return introducation
            
def enORGIllegalRevise(tokens,tags):
    i = 0
    while i < len(tokens):
        # print(i)
        if tags[i] in ["ORG","WORK-ORG","EDU-ORG","CRE-CREDIT"]:
            element,tg,end = get_element(tokens, tags, i, tags[i])
            # 第一种情况，如果是中文字符串删除
            # 第二种情况，如果只是英文停用词或者中英文标点，删除    
            # 第三种情况，字母数量小于等于2 删除        
            if not str_is_english(element) or isStopWordAndPunction(element) or getAlphbetNumber(element) < 3:
                tags = deleteTag(tags,i,end,"清空tg")
            i = end

        else:
            i = i + 1
    return tags      

# 判断给定字符串是不是英文停用词或者标点符号
def isStopWordAndPunction(word):
    stop_words = ['i', 'me', 'my', 'myself', 'we', 'our', 'ours', 'ourselves', 'you', 'your', 'yours', 'yourself', 'yourselves', 'he', 'him', 'his', 'himself', 'she', 'her', 'hers', 'herself', 'it', 'its', 'itself', 'they', 'them', 'their', 'theirs', 'themselves', 'wht', 'which', 'who', 'whom', 'this', 'tht', 'these', 'those', 'm', 'is', 're', 'ws', 'were', 'be', 'been', 'being', 'hve', 'hs', 'hd', 'hving', 'do', 'does', 'did', 'doing', '', 'n', 'the', 'and', 'but', 'if', 'or', 'becuse', 's', 'until', 'while', 'of', 't', 'by', 'for', 'with', 'bout', 'ginst', 'between', 'into', 'through', 'during', 'before', 'fter', 'bove', 'below', 'to', 'from', 'up', 'down', 'in', 'out', 'on', 'off', 'over', 'under', 'gin', 'further', 'then', 'once', 'here', 'there', 'when', 'where', 'why', 'how', 'll', 'ny', 'both', 'ech', 'few', 'more', 'most', 'other', 'some', 'such', 'no', 'nor', 'not', 'only', 'own', 'sme', 'so', 'thn', 'too', 'very', 's', 't', 'cn', 'will', 'just', 'don', 'should', 'now', 'd', 'll', 'm', 'o', 're', 've', 'y', 'in', 'ren', 'couldn', 'didn', 'doesn', 'hdn', 'hsn', 'hven', 'isn', 'm', 'mightn', 'mustn', 'needn', 'shn', 'shouldn', 'wsn', 'weren', 'won', 'wouldn']
    if word.lower() in stop_words:
        return True
    elif char_is_punctuation(word):
        return True
    else:
        return False 

# 统计字符串中字母数量
def getAlphbetNumber(element):
    num = 0 
    for s in element:
        if char_is_alphabet(s):
            num += 1
    return num

# 传入一个text 和 一个要查找的子串list 找到在字符串中所有出现子串位置的最右边
def get_last_index(text,element_list):
    mx_index = -1
    for element in element_list:
        index = text.rfind(element)
        # print(element,index)
        if index != -1 and index + len(element) > mx_index :
            # print(mx_index)
            mx_index = index + len(element)
    return mx_index

# 修正一些预测错误的情况
# 1.work-org修改原则
# 1.1被edu包含的work_org改为edu-org
# 1.2前后都不是work的org改为org
# 2.edu-org修改原则
# 2.1被work包含的edu-org改为work-org
# 3.org修改原则
# 3.1接下来是edu-degree的org改为edu-org
# 4.WORK-STRTTIME修改原则
# 4.1被edu包含的WORK-STRTTIME改为EDU-STRTYER

def basic_edu_work_credit_alignment(introducation,tags,debug):
    lstTg = tags[0]
    lstTgIndex = 0
    i = 1
    while i < len(tags):
        # print(i)
        # 如果tg为空直接跳过
        if tags[i] == "":
            i += 1
        # 被edu包含的WORK-ORG改为EDU-ORG,前后都不是work-ORG改为ORG
        elif tags[i] == "WORK-ORG":
            begin_index,end_index = get_begin_index_nd_end_index(tags,i)
            nextTg,nextTgIndex = get_next_tg(tags,end_index + 1,tags[i])
            if debug:
                print(begin_index,end_index,nextTg,nextTgIndex)
            # 第一种情况前后都是edu改为edu
            if lstTg.startswith("EDU-") and nextTg.startswith("EDU-"):
                # modifyTg方法传入的右边界应该是要修正边界+1
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"EDU-ORG","错位修正")
            # 第二种情况 如果work-org紧接着edu-degree 改为edu-org
            elif nextTg.startswith("EDU") and nextTgIndex == end_index + 1:
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"EDU-ORG","错位修正")
            # 第三种情况前后都不是work改为org
            elif not lstTg.startswith("WORK-") and not nextTg.startswith("WORK-"):
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"ORG","错位修正")
            # 如果上面三格内是edu-endyer，那么work-ORG改为edu-ORG
            elif lstTg.startswith("EDU-ENDYER") and begin_index - lstTgIndex < 3:
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"EDU-ORG","错位修正")
            
            # 将目前的标签定位上一次标签
            lstTg = tags[i]
            # 上一次标签结束地方为end_index
            lstTgIndex = end_index
            # i定位下一个标签的起始地方
            i = nextTgIndex
        # 1.被work包含的EDU-ORG改为WORK-ORG
        # 2.紧跟着work-deprt的改为work-org
        # 3.两边都不是edu的edu-org删掉
        elif tags[i] == "EDU-ORG":
            begin_index,end_index = get_begin_index_nd_end_index(tags,i)
            nextTg,nextTgIndex = get_next_tg(tags,end_index + 1,tags[i])
            if debug:
                print(begin_index,end_index,nextTg,nextTgIndex)
            if lstTg.startswith("WORK-") and nextTg.startswith("WORK-"):
                # modifyTg方法传入的右边界应该是要修正边界+1
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"WORK-ORG","错位修正")
            elif nextTgIndex - end_index == 1 and nextTg.startswith("WORK-DEPRT"):
                # modifyTg方法传入的右边界应该是要修正边界+1
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"WORK-ORG","错位修正") 
            elif nextTgIndex - end_index == 1 and nextTg.startswith("TITLE"):
                # modifyTg方法传入的右边界应该是要修正边界+1
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"ORG","错位修正")                       
            elif not lstTg.startswith("EDU-") and not nextTg.startswith("EDU-"):
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"ORG","错位修正")
            # 将目前的标签定位上一次标签
            lstTg = tags[i]
            # 上一次标签结束地方为end_index
            lstTgIndex = end_index
            # i定位下一个标签的起始地方
            i = nextTgIndex
        # 被edu包含的WORK-STRTTIME改为EDU-ENDYER，辖三个内是edu改为edu-endyer,下三格内是cre打头的改为CRE-year
        elif tags[i] == "WORK-STRTTIME":
            begin_index,end_index = get_begin_index_nd_end_index(tags,i)
            nextTg,nextTgIndex = get_next_tg(tags,end_index + 1,tags[i])
            if debug:
                print(begin_index,end_index,nextTg,nextTgIndex)
            # 第一种情况,上一个tg是EDU打头，下一个tg是EDU-ENDYER，那么tg改为EDU-STRTYER
            if lstTg.startswith("EDU-") and nextTg.startswith("EDU-ENDYER"):
                # modifyTg方法传入的右边界应该是要修正边界+1
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"EDU-STRTYER","错位修正")
            # 第二种情况,上一个tg是EDU-STRTYER,下一个是EDU，那么tg改为EDU-ENDYER
            elif lstTg.startswith("EDU-STRTYER") and nextTg.startswith("EDU-"):
                # modifyTg方法传入的右边界应该是要修正边界+1
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"EDU-ENDYER","错位修正")
            # 接下来三格以内是edu的改为edu-endyer
            elif nextTgIndex - end_index < 3 and nextTg.startswith("EDU-"):
                # modifyTg方法传入的右边界应该是要修正边界+1
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"EDU-ENDYER","错位修正")
            elif nextTgIndex - end_index < 3 and nextTg.startswith("CRE-"):
                # modifyTg方法传入的右边界应该是要修正边界+1
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"CRE-year","错位修正")
            # 将目前的标签定位上一次标签
            lstTg = tags[i]
            # 上一次标签结束地方为end_index
            lstTgIndex = end_index
            # i定位下一个标签的起始地方
            i = nextTgIndex
        # 接下来是一格是EDU的ORG改为EDU-ORG 可能是edu-degree 也可能是edu-deprtment
        elif tags[i] == "ORG":
            begin_index,end_index = get_begin_index_nd_end_index(tags,i)
            nextTg,nextTgIndex = get_next_tg(tags,end_index + 1,tags[i])
            if debug:
                print(begin_index,end_index,nextTg,nextTgIndex)
            if lstTg.startswith("EDU-") and nextTg.startswith("EDU-"):
                # modifyTg方法传入的右边界应该是要修正边界+1
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"EDU-ORG","错位修正")
            elif lstTg.startswith("WORK-") and nextTg.startswith("WORK-"):
                # modifyTg方法传入的右边界应该是要修正边界+1
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"WORK-ORG","错位修正")
            elif nextTgIndex - end_index == 1 and nextTg.startswith("EDU"):
                # modifyTg方法传入的右边界应该是要修正边界+1
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"EDU-ORG","错位修正")
            
            # 将目前的标签定位上一次标签
            lstTg = tags[i]
            # 上一次标签结束地方为end_index
            lstTgIndex = end_index
            # i定位下一个标签的起始地方
            i = nextTgIndex
        # 加入规则5.1针对所有edu-degree，如果前后都没edu就清空
        # 加入规则5.2针对所有edu-degree，如果后面是导师二字就清空
        elif tags[i] == "EDU-DEGREE":
            begin_index,end_index = get_begin_index_nd_end_index(tags,i)
            nextTg,nextTgIndex = get_next_tg(tags,end_index + 1,tags[i])
            if debug:
                print(begin_index,end_index,nextTg,nextTgIndex)
            if not lstTg.startswith("EDU-") and not nextTg.startswith("EDU-"):
                # modifyTg方法传入的右边界应该是要修正边界+1
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"","错位修正")
            elif introducation[end_index+1:end_index+3] == "导师":
                # modifyTg方法传入的右边界应该是要修正边界+1
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"","错位修正")       
            # 将目前的标签定位上一次标签
            lstTg = tags[i]
            # 上一次标签结束地方为end_index
            lstTgIndex = end_index
            # i定位下一个标签的起始地方
            i = nextTgIndex
        # 加入规则6.1针对所有的work-deprt，如果前一位置是org则修改为deprt
        elif tags[i] == "WORK-DEPRT":
            begin_index,end_index = get_begin_index_nd_end_index(tags,i)
            nextTg,nextTgIndex = get_next_tg(tags,end_index + 1,tags[i])
            if debug:
                print(begin_index,end_index,nextTg,nextTgIndex)
            if lstTg.startswith("ORG") and i - lstTgIndex < 3:
                # modifyTg方法传入的右边界应该是要修正边界+1
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"DEPRT","错位修正")            
            # 将目前的标签定位上一次标签
            lstTg = tags[i]
            # 上一次标签结束地方为end_index
            lstTgIndex = end_index
            # i定位下一个标签的起始地方
            i = nextTgIndex
        # 加入规则7.1针对所有work-post 如果前一位置是deprt则修改为post
        elif tags[i] == "WORK-POST":
            begin_index,end_index = get_begin_index_nd_end_index(tags,i)
            nextTg,nextTgIndex = get_next_tg(tags,end_index + 1,tags[i])
            if debug:
                print(begin_index,end_index,nextTg,nextTgIndex)
            if lstTg.startswith("DEPRT") and i - lstTgIndex < 3:
                # modifyTg方法传入的右边界应该是要修正边界+1
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"POST","错位修正")
            elif lstTg.startswith("ORG") and i -lstTgIndex < 3:                   
                # modifyTg方法传入的右边界应该是要修正边界+1
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"POST","错位修正")             
            # 将目前的标签定位上一次标签
            lstTg = tags[i]
            # 上一次标签结束地方为end_index
            lstTgIndex = end_index
            # i定位下一个标签的起始地方
            i = nextTgIndex
        else:      
            lstTg = tags[i]
            lstTgIndex = i
            i += 1
    return tags

# 返回tg的开始位置和结束位置
def get_begin_index_nd_end_index(tags,begin_index):
    end_index = begin_index
    while end_index < len(tags) and tags[end_index] == tags[begin_index]:
        end_index += 1
    return begin_index,end_index-1

# 如果不存在下一个标签，返回nextTg为空，nextTgIndex为当前tg结束后的第一个位置
def get_next_tg(tags,begin_index,tg):
    nextTg = ""
    nextTgIndex = begin_index
    for i in range(begin_index,len(tags)):
        if tags[i] != tg and tags[i] != "":
            nextTg = tags[i]
            nextTgIndex = i
            break
    return nextTg,nextTgIndex

def print_expert_info(expertInfo):
    print("基本信息如下:")
    print(expertInfo.bsicInfo)
    print("工作经历如下:")
    for i in range(len(expertInfo.workInfos)):
        print(i,expertInfo.workInfos[i])
    print("教育经历如下:")
    for i in range(len(expertInfo.eductionInfos)):
        print(i,expertInfo.eductionInfos[i])
    print("荣誉获奖如下:")
    for i in range(len(expertInfo.creditInfos)):
        print(i,expertInfo.creditInfos[i])

#废弃的方法
# def strangeSymbolReplace(s):
#     # 用英语字母替代德语
#     s = s.replace("ö","o").replace("ü","u").replace("Ö","o").replace("ï","i").replace("ä","").replace("Å","").replace("ó","o")
#     # 去掉韩语和日语 \uC00-\uD73为匹配韩文的，其余为日文
#     s = re.sub(r'[\u3040-\u309F\u300-\u30FF\uC00-\uD73]', '', s) 
#     # 用英语字母替代法语
#     s = replceFrenchLetterWithEnglishLetter(s)
#     return s


# def replceFrenchLetterWithEnglishLetter(s):
#     s = s.replace('é', 'e')
#     s = s.replace('à', '')
#     s = s.replace('è', 'e')
#     s = s.replace('ù', 'u')
#     s = s.replace('â', 's')
#     s = s.replace('ê', 'e')
#     s = s.replace('î', 'i')
#     s = s.replace('ô', 'o')
#     s = s.replace('û', 'u')
#     s = s.replace('ç', 'c')
#     return s

# def initTags(tokens,origin_tgs):
#     tags = []
#     for i in range(len(tokens)):
#         for j in range(len(tokens[i])):
#             tags.append(origin_tgs[i])
#     return tags

# # 将第一个列表中出现的[UNK] 替换为第二个字符串中得指定值,如果多个UNK连续出现
# def replceUNK(intro_list,intro_str):
    # str_index = 0
    # list_index = 0
    # # 循环遍历列表中的所有值
    # while list_index < len(intro_list):
    #     current_token = intro_list[list_index] 
    #     # 如果当前token不是unk，那么就将str_index往前移，同时ist_index移到下一个
    #     if current_token != "[UNK]":
    #         str_index += len(current_token)            
    #         list_index += 1
    #     else:
    #         print(intro_list[list_index-1])
    #         # print(intro_list[list_index+1])
    #         print(intro_str[str_index:str_index+10])
           
    #         intro_list[list_index] = intro_str[str_index:str_index+1]
    #         str_index += 1
    #         list_index += 1
    # return intro_list,intro_str


    
    

  