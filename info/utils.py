import jieba
import xlrd
import openpyxl
import re
import string
import json
from transformers import BertTokenizer
from zhon.hanzi import punctuation
from tqdm import tqdm
# 判断一个字符是不是英文字符
def char_is_alphabet(uchar):
    """判断一个unicode是否是英文字母"""
    if (u'\u0041' <= uchar <= u'\u005a') or (u'\u0061' <= uchar <= u'\u007a'):
        return True
    else:
        return False

# 判断一个字符是不是数字
def char_is_number(uchar):
    """判断一个unicode是否是数字"""
    if u'\u0030' <= uchar <= u'\u0039':
        return True
    else:
        return False

# 判断一个字符是不是汉字
def char_is_chinese(uchar):
    """判断一个unicode是否是汉字"""
    if  u'\u4e00' <= uchar <= u'\u9fa5':
        return True
    else:
        return False

# 判断一个字符是不是括号
def char_is_parentheses(uchar):
    if uchar in ["(",")","（","）"]:
        return True
    else:
        return False

# 判断一个字符是不是单引号或者双引号
def char_is_quotation(uchar):
    if uchar in ["”","“","‘","’","'","\""]:
        return True
    else:
        return False

# 判断一个字符是不是空格
def char_is_space(uchar):
    """判断一个unicode是否是空字符串（包括空格，回车，tab）"""
    space = [u'\u0020', u'\u000A', u'\u000D', u'\u0009']
    if uchar in space:
        return True
    else:
        return False

# 判断一个字符是不是标点符号
def char_is_punctuation(uchar):
    if uchar in string.punctuation or uchar in punctuation:
        return True
    else:
        return False

# 判断一个字符是否非汉字，数字，空字符和英文字符还不是标点符号
def char_is_other(uchar):
    """判断是否非汉字，数字，空字符和英文字符"""
    if not (char_is_space(uchar) or char_is_chinese(uchar) or char_is_number(uchar) or char_is_alphabet(uchar) or char_is_punctuation(uchar) ):
        return True
    else:
        return False

# 判断一个字符串是不是英文字符串，如果一个字符串英文字符长度是中文字符长度5倍就认为他是英文字符串，并且英文字符长度不为0
def str_is_english(str):
    e_num = 0
    c_num =0
    for char in str:
        if char_is_alphabet(char):
            e_num += 1
        elif char_is_chinese(char):
            c_num += 1
    if e_num > 5*c_num:
        return True
    else:
        return False

# 判断一个字符串是不是中文字符串，如果一个字符串英文字符长度小于中文字符长度5倍就认为他是中文字符串，并且中文字符长度不为0
def str_is_chinese(str):
    e_num = 0
    c_num =0
    for char in str:
        if char_is_alphabet(char):
            e_num += 1
        elif char_is_chinese(char):
            c_num += 1
    if  5 *c_num >= e_num:
        return True
    else:
        return False

# 判断一个字符串是否不含英文字母
def str_without_alphabet(str):
    for char in str:
        if char_is_alphabet(char):
            return False
    return True

# 判断一个字符串是否是int
def str_is_number(str):
    try:
        int(str)
        return True
    except ValueError:
        pass

# 得到一个字符串中字母的数量
def getStringAlphabetNumber(str):
    num = 0
    for char in str:
        if char_is_alphabet(char):
            num += 1
    return num

# 得到一个字符串中汉字的数量
def getStringChineseNumber(str):
    num = 0
    for char in str:
        if char_is_chinese(char):
            num += 1
    return num

# 判断专家介绍为英文或中文，同时将长介绍修正
def introducation_is_english_or_chinese(introducation):
    language = ""
    if str_is_chinese(introducation[:100]) or getStringChineseNumber(introducation) > 50 :
        language = "chinese"
    else:
        language = "english"
    if len(introducation) > 4000:
        introducation = introducation[0:4000]
    return language,introducation

# 通过xlrd工具读取所有专家数据，适用于xls后缀文件,第二个参数为想要读取专家数量，默认为全部
def readInputDataByXlrd(filename,experts_num=-1):
    xlsx = xlrd.open_workbook(filename)
    sheet = xlsx.sheets()[0]
    rows = sheet.nrows
    cols = sheet.ncols

    # 如果没有指定专家数量,则返回所有专家
    if experts_num == -1:        
        experts_num = rows   

    key_list = []
    # 获得key_list id 名字 所在机构 简介
    for i in range(cols):
        key_list.append(sheet.cell(0, i).value)
    # 获得experts
    experts = {}
    for i in range(1, experts_num):
    # for i in tqdm(range(1, experts_num)):
        expert_id =  str(sheet.cell(i, 0).value)
        expert = {}
        # print(sheet.cell(i, 0).value)
        introducation = ""
        # # 获得专家 id 名字 所在机构信息
        
        for j in range(0, 3):
            expert[key_list[j]] = str(sheet.cell(i, j).value)

        # # 专家的简介信息需要合并第四列到第九列
        for j in range(3, cols):
            if type(sheet.cell(i, j).value) == float:
                introducation += str(int(sheet.cell(i, j).value))
            else:
                introducation += str(sheet.cell(i, j).value)
        # 用正则表达式删除html标签，之后删掉&nbsp,换行符和制表符
        expert[key_list[3]] = reviseIntroducation(introducation)
        expert["原始简介"] = introducation
        experts[expert_id] = expert
    return experts

# 通过openpyxl工具读取所有专家数据，适用于xlsx后缀文件，第二个参数为想要读取专家数量，默认为全部
def readInputDataByOpenpyxl(filename,experts_num=-1):
    # 读取excel
    workbook = openpyxl.load_workbook(filename)
    # 获取所有sheet的名字
    sheetnames = workbook.sheetnames 
    # 加载第一个sheet
    sheet = workbook[sheetnames[0]] 
    # 获取sheet的行数和列数
    sheet_rows = sheet.max_row
    sheet_cols = sheet.max_column

    # 如果没有指定专家数量,则返回所有专家
    if experts_num == -1:        
        experts_num = sheet_rows
    
    # 获得key_list id 名字 所在机构 简介 openpyxl下标从1开始
    key_list = []  
    for i in range(1,sheet_cols+1):
        key_list.append(sheet.cell(1, i).value)
    
    # 获得experts
    experts = {}
    for i in tqdm(range(2, experts_num + 1)):
        expert_id =  str(sheet.cell(i, 1).value)
        expert = {}
        # print(sheet.cell(i, 0).value)
        introducation = ""
        # # 获得专家 id 名字 所在机构信息        
        for j in range(1, 4):
            expert[key_list[j]] = str(sheet.cell(i, j).value)
        # # 专家的简介信息需要合并第四列到第九列
        for j in range(4,sheet_cols + 1):
            if type(sheet.cell(i, j).value) == float:
                introducation += str(int(sheet.cell(i, j).value))
            elif str(sheet.cell(i, j).value) != "None":
                introducation += str(sheet.cell(i, j).value)
        # 用正则表达式删除html标签，之后删掉&nbsp,换行符和制表符
        expert[key_list[3]] = reviseIntroducation(introducation)
        expert["原始简介"] = introducation
        experts[expert_id] = expert
    return experts

# 通过json文件读取所有专家数据
def readInputDataByJson(filename):
    with open(filename,'r',encoding='utf8')as f:
        results = json.load(f)
    return results

# 读取预测结果
def readPredictResult(filename):
    with open(filename,'r',encoding='utf8')as f:
        results = json.load(f)
    return results

# 读取ner结果
def getORGParticiple(filename):
    with open(filename,'r',encoding='utf8')as f:
        results = json.load(f)
    return results

# 测试数据
def getTestData():
    experts = {}
    expert = {}
    expert['id'] = "1"
    expert['简介'] = "李冉，男，复旦大学教授，1979年10月1日生，国籍：中国，汉族，江苏南京人，担任马克思主义学院院长；马克思主义研究院副院长，" \
                             "研究方向：马克思主义，学科领域，政治学，通讯地址光华楼101，邮编：201101，电话456789，手机123456" \
                             "传真88880，邮箱88888@qq.com，个人网页www.baidu.com" \
                   "教育经历：南京大学本科研究生连读 1995年进入清华大学机械工程系修读液压传动及气动博士 2000年毕业" \
                   "获奖经历 2010，2011连续两年获得长江学者称号，2012年获得世界级别的由哈佛大学在美国颁发的持续到2014年的教学成果奖金奖" \
                   "工作经历 2016年在北京大学国际关系学院任院长，同时在物理学院任副院长"
    experts[expert['id']] = expert

    # marked_experts = {}
    # marked_expert = {}
    # marked_expert['cname'] = "李冉"
    # marked_expert['ename'] = ""
    # marked_expert['corganization'] = "复旦大学"
    # marked_expert['eorganization'] = ""
    # marked_expert['title'] = "教授"
    # marked_expert['year'] = "1979"
    # marked_expert['month'] = "10"
    # marked_expert['day'] = "1"
    # marked_expert['sex'] = "男"
    # marked_expert['nationality'] = "中国"
    # marked_expert['nation'] = "汉族"
    # marked_expert['nativePlace'] = "江苏南京"
    # marked_expert['depart'] = "马克思主义学院；马克思主义研究院"
    # marked_expert['post'] = "马克思主义学院院长；马克思主义研究院副院长"
    # marked_expert["interest"] = "马克思主义"
    # marked_expert["major"] = "政治学"
    # marked_expert["address"] = "光华楼101"
    # marked_expert["code"] = "201101"
    # marked_expert["phone"] = "456789"
    # marked_expert["telephone"] = "123456"
    # marked_expert["fax"] = "88880"
    # marked_expert["mail"] = "88888@qq.com"
    # marked_expert["website"] = "www.baidu.com"

    # marked_expert["education"] = []
    # marked_expert["education"].append({"cschool": "南京大学", "education": "本科"})
    # marked_expert["education"].append({"cschool": "南京大学", "education": "研究生"})
    # marked_expert["education"].append({"cschool": "清华大学", "startYear": "1995","endYear":"2000","department":"机械工程系",
    #                                     "major":"液压传动及气动","degree":"博士"})
    # marked_expert["credit"] = []
    # marked_expert["credit"].append({"credit": "长江学者", "year": "2010"})
    # marked_expert["credit"].append({"credit": "长江学者", "year": "2011"})
    # marked_expert["credit"].append({"credit": "教学成果奖", "year": "2012","level":"世界级","organization":"哈佛大学",
    #                                 "place":"美国","endYear":"2014","grade":"金奖"})
    # marked_expert["work"] = []
    # marked_expert["work"].append({"corganization": "北京大学", "depart":"国际关系学院","startTime": "2016","post":"国际关系学院院长"})
    # marked_expert["work"].append({"corganization": "北京大学", "depart":"物理学院","startTime": "2016","post":"物理学院副院长"})
    # marked_expert["credit"].append({"credit": "教学成果奖", "year": "2012", "level": "世界级", "organization": "哈佛大学",
    #                                 "place": "美国", "endYear": "2014", "grade": "金奖"})
    # marked_experts[expert['id']] = marked_expert
    return experts

# 字符串如果小于指定长度就清空
def check_len_more_than_threshold(str,length):
    if len(str) < length:
        str = ""
    return str

# 修正时间表示
def reviseTime(str_time):
    # 如果字符串中有年和月
    if "年" in str_time and "月" in str_time:
        index1 = str_time.find("年")
        index2 = str_time.find("月")
        # 找到年和月
        year = str_time[0:index1]
        month = str_time[index1+1:index2]
        # 如果月份以0打头删掉
        if(month.startswith("0")):
            month = month[1:]
        # 年份是数字，月份是数字，且年份大于1800小于2025 月份 大于0小于13 才返回
        if str_is_number(year) and str_is_number(month) and int(year) > 1800 and int(year) < 2025 and int(month) > 0 and  int(month) < 13:
                return year + "年" + month + "月"
        else:
            return ""
    # 如果字符串只有年
    elif "年" in str_time:
        index1 = str_time.find("年")
        # 找到年
        year = str_time[0:index1]
        # 如果年后面还有东西则代表可能有月信息，提取月信息
        if len(year) == 4 and len(str_time) > index1  + 1 :
            month = str_time[index1+1]
            # 如果月信息和年信息格式都正确
            if str_is_number(year) and str_is_number(month) and int(year) > 1800 and int(year) < 2050 and int(month) > 0 and  int(month) < 13:
                return year + "年" + month + "月"
            # 否则年信息格式正确，只输出年
            elif str_is_number(year) and int(year) > 1800 and int(year) < 2050:
                return year + "年"
            else:
                return ""
        # 否则只有年信息，保证年份为1800-2050之间的数字即可
        elif str_is_number(year) and int(year) > 1800 and int(year) < 2050 :
            return year + "年"
        else:
            return ""
    # 如果字符串是单纯的四位数字
    elif str_is_number(str_time) and int(str_time) > 1800 and int(str_time) < 2050 :
        return str_time + "年"
    # 如果字符串以点分隔年月
    elif "." in str_time:
        arr = str_time.split(".")
        if(arr[1].startswith("0")):
            arr[1] = arr[1][1:]
        year = arr[0]
        month =arr[1]
         # 如果月信息和年信息格式都正确
        if str_is_number(year) and str_is_number(month) and int(year) > 1800 and int(year) < 2050 and int(month) > 0 and int(month) < 13:
                return year + "年" + month + "月"
        # 否则年信息格式正确，只输出年
        elif str_is_number(year) and int(year) > 1800 and int(year) < 2050:
                return year + "年"
        else:
            return ""
    else:
        return ""

# 检查结束时间是否大于开始时间
def checkTime(startTime,endTime,interval=0):

    # 如果有一个时间为空直接返回
    if startTime == "" or endTime == "":
        return endTime
    
    # 找到年份，
    startYear = int(startTime[0:startTime.find("年")])
    endYear = int(endTime[0:endTime.find("年")])
    if endYear - startYear >= interval:
        return endTime
    elif endYear < startYear or interval > 0 :
        return ""
    
    # 年份相同的情况下，且interval参数为0的情况 如果都有月份
    if "月" in startTime and "月" in endTime:
        startMonth = int(startTime[startTime.find("年") +1 :startTime.find("月")])
        endMonth = int(endTime[endTime.find("年") +1 :endTime.find("月")])
        if endMonth > startMonth:
            return endTime
        else:
            return ""
    # 如果都没月份，则判断成立
    if "月" not in startTime and "月" not in endTime:
        return endTime
    
    # 如果一个有月份一个没月份，默认出错，endTime清零
    return ""

# decoder时候获取element 
def get_element(tokens,tags,i,tag):
    start = i
    end = i + 1
    res = tokens[start]
    while end < len(tokens) and tags[end] == tag:
        res += tokens[end]
        end = end + 1
    return res,tag,end

# 相邻合并
def mergeAdjacentTag(introducation,tags,distance,debug=False):
    lastIndex = 0
    lastTag = tags[0]
    for i in range(len(tags)):
        # 中间只有中文,英文字母，数字,空格，或者括号，或者-才合并
        if tags[i] != "" and tags[i] == lastTag and i - lastIndex <= distance and i - lastIndex > 1 and checkMergeCondition(lastTag,introducation,tags,lastIndex+1,i):
            for j in range(lastIndex+1,i):
                if debug:
                    print("相邻合并",j)
                tags[j] = lastTag                  
            lastIndex = i
            lastTag = tags[i]
        elif tags[i] != "":
            lastIndex = i
            lastTag = tags[i]
    return tags

# 检查是否满足合并条件
def checkMergeCondition(tag,introducation,tags,begin,end):
    for i in range(begin,end):        
        char = introducation[i]
        if char_is_chinese(char) :
            continue
        elif char_is_alphabet(char):        
            continue
        elif char_is_number(char):
            continue
        elif char_is_space(char):
            continue
        elif char_is_parentheses(char):
            continue
        elif char_is_quotation(char):
            continue
        # 如果要补充的tag为NAME并且char为.的话
        elif tag == "NAME" and char == ".":
            continue
        elif char == "—":
            continue
        else:
            return False
    return True

# 修复中英文括号，引号，书名号不对齐相关问题，同时
def reviseSymbolPair(s):
    s = s.strip()
    # 如果字符串是单个符号，那么不补全，直接删除
    if len(s) == 1 and s in [")","）","(","（","“","”","《","》"]:
        s = ""
    # 删掉最后多余符号
    if len(s) > 0 and s[0] in [")","）","》","\""]:
        s = s[1:]
    if len(s) > 0 and s[-1] in ["(","（","《","\""]:
        s = s[:-1]
        
    # 补充括号
    if s.count("(") - s.count(")") == 1:
        s = s +")"
    elif s.count("（") - s.count("）") == 1:
        s = s +"）"
    elif s.count("(") - s.count(")") == -1:
        s = "(" + s 
    elif s.count("（") - s.count("）") == -1:
        s = "（" + s 
    
    # 补充引号
    if s.count("“") - s.count("”")  == 1:
        s = s + "”"
    elif s.count("“") - s.count("”") == -1:               
        s = "“" + s
    

    
    # 补充书名号
    if s.count("《") - s.count("》")  == 1:
        s = s + "》"
    elif s.count("《") - s.count("》") == -1:
        s = "《" + s


    return s.strip() 


# 修复一些乱七八糟的前缀和后缀 后两个参数可选，调用者可以补充要修复的前缀和后缀，比如学科领域要删掉后缀家
def revisePrefixAndSuffix(s,prefix_dict_supplement=[],suffix_dict_supplement=[]):
    s = s.strip()
    delete_prefix_dict =    [",","，", 
                              ".","。",
                              ":","：",
                              "、","/",
                              "-","——",
                              "&","＆",
                              "'",
                              "及","与","于","和","的"]
    # "任"
    delete_prefix_dict.extend(prefix_dict_supplement)
    
    delete_suffix_dict =    [",","，", 
                              ".","。",
                              ":","：",
                              "、","/",
                              "-","——",
                              "&","＆",
                              "'",
                              "及","与","于","和","的"]
    # ,"任"
    delete_suffix_dict.extend(suffix_dict_supplement)
        

    # 删除指定前缀
    for prefix in delete_prefix_dict:
        if s.startswith(prefix):
            s= s[len(prefix):]
            break
        
    # 删除指定后缀
    for suffix in delete_suffix_dict:
        if s.endswith(suffix):
            s = s[:-len(suffix)]
            break
    return s.strip()

# 检查从begin 到 end 中的token有没有出现句号
def checkFullStopInTokens(tokens,begin,end):
    for i in range(begin,end):
        if tokens[i] in [".","。"]:
            return True
    return False

# 检查从begin到end中的token有没有出现分号
def checkSemicolonInTokens(tokens,begin,end):
    for i in range(begin,end):
        if tokens[i] in [";","；"]:
            return True
    return False

# 修复荣誉获奖 英文奖项名去掉stopword长度为1的清空
def reviseEnglishItem(s):
    s_copy = s
    s = s.strip()
    s_arr= s.split(" ")
 
    stop_words = ['i', 'me', 'my', 'myself', 'we', 'our', 'ours', 'ourselves', 'you', 'your', 'yours', 'yourself', 'yourselves', 'he', 'him', 'his', 'himself', 'she', 'her', 'hers', 'herself', 'it', 'its', 'itself', 'they', 'them', 'their', 'theirs', 'themselves', 'what', 'which', 'who', 'whom', 'this', 'that', 'these', 'those', 'am', 'is', 'are', 'was', 'were', 'be', 'been', 'being', 'have', 'has', 'had', 'having', 'do', 'does', 'did', 'doing', 'a', 'an', 'the', 'and', 'but', 'if', 'or', 'because', 'as', 'until', 'while', 'of', 'at', 'by', 'for', 'with', 'about', 'against', 'between', 'into', 'through', 'during', 'before', 'after', 'above', 'below', 'to', 'from', 'up', 'down', 'in', 'out', 'on', 'off', 'over', 'under', 'again', 'further', 'then', 'once', 'here', 'there', 'when', 'where', 'why', 'how', 'all', 'any', 'both', 'each', 'few', 'more', 'most', 'other', 'some', 'such', 'no', 'nor', 'not', 'only', 'own', 'same', 'so', 'than', 'too', 'very', 's', 't', 'can', 'will', 'just', 'don', 'should', 'now', 'd', 'll', 'm', 'o', 're', 've', 'y', 'ain', 'aren', 'couldn', 'didn', 'doesn', 'hadn', 'hasn', 'haven', 'isn', 'ma', 'mightn', 'mustn', 'needn', 'shan', 'shouldn', 'wasn', 'weren', 'won', 'wouldn']
    for word in stop_words:
        if len(s_arr) > 0 and s_arr[0].lower() == word:
            s_arr[0] = "" 
        if len(s_arr) > 0 and s_arr[-1].lower() == word:
            s_arr[-1] = ""
    s = " ".join(s_arr)
    s = s.strip()
    if len(s.split(" ")) == 1:
        s = ""
    else:
        s = s_copy
    return s

# 中文分词边界补全
def cnBoundaryCompletion(introducation,tags,debug):
    # doc = zh_nlp(introducation[0:100])
    begin_indexz,end_indexz = getCNParticiple(introducation)
    # print(begin_indexz)
    # print(end_indexz)
    # 得到所有左边界和右边界
    # for sentence in doc.sentences:
    #     for token in sentence.tokens:
    #         # print(token)
    #         begin_indexz.append(token.start_char)
    #         end_indexz.append(token.end_char)
    # print(begin_indexz)
    # print(end_indexz)

    # 循环遍历第二个到倒数第二个tag
    for i in range(1,len(tags)-1):
        # 第一种情况如果tag不等空且上一个tag不等于这个tag，代表左边界的情况出现了
        if tags[i] != "" and tags[i-1] != tags[i]:
            # 如果i不在中文分词左边界，找到列表中小于i最大的左边界
            left_boundary = getLeftBoundary(i,begin_indexz)
            # 修改tag
            tags = modifyTag(debug,introducation,tags,left_boundary,i,tags[i],"左边界")                     
        # 如果tag不等于空，且下一个tag不等于这个tag，代表右边界的情况出现了
        if tags[i] != "" and tags[i+1] != tags[i]:
            # 如果i+1不在中文分词右边界，找到列表中大于i+1最小的右边界      
            right_boundary = getRightBoundary(i+1,end_indexz)
            tags = modifyTag(debug,introducation,tags,i+1,right_boundary,tags[i],"右边界")

    return tags

# 英文chunking边界补全
def enBoundaryCompletion(introducation,tags,en_chunking_nlp):
    # doc = zh_nlp(introducation[0:100])
    chunking_index_list = getENParticiple(introducation,en_chunking_nlp)
    # print(chunking_index_list)
    for i in range(1,len(tags)-1):
        if judgeWhetherInORGPaticiple(chunking_index_list,i) != -1:           
            left_boundary,right_boundary  =chunking_index_list[judgeWhetherInORGPaticiple(chunking_index_list,i)]
            modifyTag(False,introducation,tags,left_boundary,right_boundary,tags[i],"ner修改",["TITLE","POST","MAJOR","DEPART","WORK-DEPART","EDU-DEPARTMENT"])
            i = right_boundary - 1        
    return tags

# 中文组织ner边界补全  
def cnORGBoundaryRevise(introducation,tags,org_index_list,debug):
    for i in range(1,len(tags)-1):
        # 先做stanza ner 边界补全
        if tags[i] in ["ORG","WORK-ORG","EDU-ORG"] and judgeWhetherInORGPaticiple(org_index_list,i) != -1:           
            left_boundary,right_boundary = org_index_list[judgeWhetherInORGPaticiple(org_index_list,i)]
            modifyTag(debug,introducation,tags,left_boundary,right_boundary,tags[i],"ner修改",["TITLE","POST","MAJOR","DEPART","WORK-DEPART","EDU-DEPARTMENT"])
            i = right_boundary - 1
    return tags

# 英文组织ner边界补全
def enORGBoundaryRevise(introducation,tags,org_index_list):
    i = 1
    while i < len(tags) - 1:
        if tags[i] in ["ORG","WORK-ORG","EDU-ORG"] and judgeWhetherInORGPaticiple(org_index_list,i) != -1:           
            left_boundary,right_boundary = org_index_list[judgeWhetherInORGPaticiple(org_index_list,i)]
            modifyTag(False,introducation,tags,left_boundary,right_boundary,tags[i],"ner修改",["TITLE","POST","MAJOR","DEPART","WORK-DEPART","EDU-DEPARTMENT"])
            i = right_boundary  
        else:
            i = i + 1       
    return tags

# 修改begin到end的tag
def modifyTag(debug,introducation,tags,begin,end,tag,hint,tag_list=[]):
    # 先检查要修改的tag有没有包含tag_list中的tag，如果包含直接整个不修改
    flag = False
    for i in range(begin,end):
        if tags[i] in tag_list:
            flag = True
            break
    # print(flag)
    if  not flag:
        for i in range(begin,end):
            tags[i] = tag
            if debug:
                print("modify"+hint,i,tags[i])
    return tags

# 得到左边界
def getLeftBoundary(index,indexz):
    if index not in indexz:
        # 特殊情况index比indexz最后一个index还大，那说明左边界就是最后一个index
        if index > indexz[-1]:
            return indexz[-1]
        # 其他情况遍历indexz中所有index，如果当前小，下一个大则代表为当前
        for i in range(len(indexz)-1):            
            if indexz[i] < index and indexz[i+1]> index:
                return indexz[i]
    return index

# 得到右边界
def getRightBoundary(index,indexz):
    if index not in indexz:
        # 特殊情况 index比indexz中第一个index还小，那说明右边界就是第一个index
        if index < indexz[0]:
            return indexz[0]
        for i in range(len(indexz)-1):
            if indexz[i] < index and indexz[i+1]> index:
                return indexz[i+1]
    return index

# 将指定begin到end index的tag清空，修正英文组织和荣誉奖励时候用到
def deleteTag(tags,begin,end,hint):
    for i in range(begin,end):
        tags[i] = ""
    return tags

# 初始化tag 英文中文都用这个方法
def initTags(key,introducation,tokens,origin_tags):
    index = 0
    tags = []
    i = 0
    while i < len(tokens):
        # [UNK] 对应的并不一定是一位
        if(tokens[i] == "[UNK]"):            
            j = 1
            # 可能存在连续的[UNK],找到后面不是[UNK]的词
            while (i+j) < len(tokens) and tokens[i+j] == "[UNK]":
                j = j + 1
            # 如果找到了不是UNK的token
            if (i+j) < len(tokens):
                #从index 开始找这个token所在的位置
                end_index = introducation.lower().find(tokens[i+j],index)
                for k in range(index,end_index):
                    tags.append("")
                # print(index,end_index,tokens[i+j])
                index = end_index
            # 说明UNK一直连到末尾            
            else:
                for k in range(index,len(introducation)):
                    # tags.append("UNK")
                    tags.append("")
                index = len(introducation)
            i = i + j        
        else:
            if(introducation[index:index+len(tokens[i])].lower() != tokens[i]):
                print(key)
                print(introducation)
                # print(len(tokens))
                print(str(index) +"\t" + str(i))
                print(str(len(tokens[i])) + "\t"  + tokens[i])
                for j in range(len(tokens[i])):
                    print(tokens[i][j])
                print(str(len(introducation[index:index+len(tokens[i])].lower())) + "\t" + introducation[index:index+len(tokens[i])].lower())
                for j in range(len(origin_tags)):
                    print(str(j) + "\t" +introducation[j] + "\t" + tokens[j] +"\t" + origin_tags[j])
                
                # print(str(index) +"\t" + str(i) + "\t" + introducation[index:index+len(tokens[i])].lower() + "\t" + tokens[i])
                exit()
            for j in range(len(tokens[i])):           
                tags.append(origin_tags[i])
            index += len(tokens[i])
            i += 1
       
        # print(index,introducation[index])
        # 一个token结束之后每有一个空格就加一个
        while index < len(introducation) and introducation[index] == " ":
            tags.append("")
            index += 1
    return tags

# 加载中文分词结果
def getCNParticiple(introducation):
    begin_indexz = []
    end_indexz = []
    jieba.load_userdict("/home/cly/poc/dict/cn_vocabulary.txt")
    seg_list = jieba.lcut(introducation, cut_all=False)
    index = 0
    for seg in seg_list:
        begin_indexz.append(index)
        index = index + len(seg)
        end_indexz.append(index)
    return begin_indexz,end_indexz

# 加载英文ORG结果
def getENORGParticiple(introducation,en_nlp):
    org_index_list = []
    doc = en_nlp(introducation)
    for sent in doc.sentences:
        for ent in sent.ents:
            if ent.type == "ORG":
                # print(ent)
                org_index_list.append((ent.start_char,ent.end_char))
    return org_index_list

# 判断是否属于org
def judgeWhetherInORGPaticiple(org_index_list,index):
    for i in range(len(org_index_list)):
        org_index = org_index_list[i]
        if index >= org_index[0] and index < org_index[1]:
            return i
    return -1

# 英文chunking的时候是组合名词短语，但注意要同一语系，以空格做分割，判断空格前和空格后是否属于同一语种
# 属于则分割，不属于则分开
def getENParticiple(introducation,en_chunking_nlp):
    end_index = 0
    chunking_index_list = []
    doc = en_chunking_nlp(introducation)
    for chunk in doc.noun_chunks:
        current_chunk_list = chunk.text.split()
        # print(current_chunk_list)
        revised_chunk_list = []
        language = " "
        for chunk_list_element in current_chunk_list:
            # 第一种情况 再识别第一个元素，判断他是中文还是英文，之后直接将结果添加进文本中
            if language == " " :
                language = "en" if str_is_english(chunk_list_element) else "cn" 
                revised_chunk_list.append(chunk_list_element)
            # 第二种情况，上一个元素为英文，当前元素也为英文，直接将当前元素加到上个元素后面         
            elif language == "en" and str_is_english(chunk_list_element):
                revised_chunk_list[-1] += " " + chunk_list_element
            # 第三种情况，上一个元素为中文，当前元素也为中文，直接将当前元素加到上个元素后面         
            elif language == "cn" and not str_is_english(chunk_list_element):
                # print(chunk_list_element)
                revised_chunk_list[-1] += " " + chunk_list_element
            # 最后一种情况，当前元素和上一个元素不一致的情况
            else:
                language = "en" if language == "cn" else "cn"
                revised_chunk_list.append(chunk_list_element)
        for chunk_element in revised_chunk_list:
            # print(chunk_element)
            begin_index = introducation.find(chunk_element,end_index)
            end_index = begin_index + len(chunk_element)
            chunking_index_list.append((begin_index,end_index))
    return chunking_index_list

# 修正个人介绍
def reviseIntroducation(introducation): 
    # 去掉html标签
    introducation = re.sub('<[^<]+?>', '', introducation)
    # 去掉换行符
    introducation = introducation.replace("\n"," ").replace("\r"," ")
    # 去掉 "\t"
    introducation = introducation.replace("\t"," ")
    # 去掉&nbsp
    introducation = introducation.replace("&nbsp;"," ")
    # 去除多余空格
    introducation = replaceMutiSpaceToOneSpace(introducation)
    # 替换一些奇怪语言
    introducation = strangeSymbolReplace(introducation)
    # 删掉两头空白
    introducation = introducation.strip()
    return introducation

# 将字符串中得多个空格替换为一个空格
def replaceMutiSpaceToOneSpace(s): 
    return " ".join(s.split())

# 删除所有空格
def deleteAllSpace(s):
    return "".join(s.split())

# 将其他语言剔除
def strangeSymbolReplace(s):
    introducation = ""
    for uchar in s:
        if not char_is_other(uchar):
            introducation += uchar
        # else:
        #     print(uchar)
    return introducation
 
# 将英文停用词剔除            
def enORGIllegalRevise(tokens,tags):
    i = 0
    while i < len(tokens):
        # print(i)
        if tags[i] in ["ORG","WORK-ORG","EDU-ORG","CRE-CREDIT"]:
            element,tag,end = get_element(tokens, tags, i, tags[i])
            # 第一种情况，如果是中文字符串删除
            # 第二种情况，如果只是英文停用词或者中英文标点，删除    
            # 第三种情况，字母数量小于等于2 删除        
            if not str_is_english(element) or isStopWordAndPuncation(element) or getAlphabetNumber(element) < 3:
                tags = deleteTag(tags,i,end,"清空tag")
            i = end

        else:
            i = i + 1
    return tags      

# 判断给定字符串是不是英文停用词或者标点符号
def isStopWordAndPuncation(word):
    stop_words = ['i', 'me', 'my', 'myself', 'we', 'our', 'ours', 'ourselves', 'you', 'your', 'yours', 'yourself', 'yourselves', 'he', 'him', 'his', 'himself', 'she', 'her', 'hers', 'herself', 'it', 'its', 'itself', 'they', 'them', 'their', 'theirs', 'themselves', 'what', 'which', 'who', 'whom', 'this', 'that', 'these', 'those', 'am', 'is', 'are', 'was', 'were', 'be', 'been', 'being', 'have', 'has', 'had', 'having', 'do', 'does', 'did', 'doing', 'a', 'an', 'the', 'and', 'but', 'if', 'or', 'because', 'as', 'until', 'while', 'of', 'at', 'by', 'for', 'with', 'about', 'against', 'between', 'into', 'through', 'during', 'before', 'after', 'above', 'below', 'to', 'from', 'up', 'down', 'in', 'out', 'on', 'off', 'over', 'under', 'again', 'further', 'then', 'once', 'here', 'there', 'when', 'where', 'why', 'how', 'all', 'any', 'both', 'each', 'few', 'more', 'most', 'other', 'some', 'such', 'no', 'nor', 'not', 'only', 'own', 'same', 'so', 'than', 'too', 'very', 's', 't', 'can', 'will', 'just', 'don', 'should', 'now', 'd', 'll', 'm', 'o', 're', 've', 'y', 'ain', 'aren', 'couldn', 'didn', 'doesn', 'hadn', 'hasn', 'haven', 'isn', 'ma', 'mightn', 'mustn', 'needn', 'shan', 'shouldn', 'wasn', 'weren', 'won', 'wouldn']
    if word.lower() in stop_words:
        return True
    elif char_is_punctuation(word):
        return True
    else:
        return False 

# 统计字符串中字母数量
def getAlphabetNumber(element):
    num = 0 
    for s in element:
        if char_is_alphabet(s):
            num += 1
    return num

# 传入一个text 和 一个要查找的子串list 找到在字符串中所有出现子串位置的最右边
def get_last_index(text,element_list):
    max_index = -1
    for element in element_list:
        index = text.rfind(element)
        # print(element,index)
        if index != -1 and index + len(element) > max_index :
            # print(max_index)
            max_index = index + len(element)
    return max_index

# 修正一些预测错误的情况
# 1.work-org修改原则
# 1.1被edu包含的work_org改为edu-org
# 1.2前后都不是work的org改为org
# 2.edu-org修改原则
# 2.1被work包含的edu-org改为work-org
# 3.org修改原则
# 3.1接下来是edu-degree的org改为edu-org
# 4.WORK-STARTTIME修改原则
# 4.1被edu包含的WORK-STARTTIME改为EDU-STARTYEAR
def basic_edu_work_credit_alignment(introducation,tags,debug):
    lastTag = tags[0]
    lastTagIndex = 0
    i = 1
    while i < len(tags):
        # print(i)
        # 如果tag为空直接跳过
        if tags[i] == "":
            i += 1
        # 被edu包含的WORK-ORG改为EDU-ORG,前后都不是work-ORG改为ORG
        elif tags[i] == "WORK-ORG":
            begin_index,end_index = get_begin_index_and_end_index(tags,i)
            nextTag,nextTagIndex = get_next_tag(tags,end_index + 1,tags[i])
            if debug:
                print(begin_index,end_index,nextTag,nextTagIndex)
            # 第一种情况前后都是edu改为edu
            if lastTag.startswith("EDU-") and nextTag.startswith("EDU-"):
                # modifyTag方法传入的右边界应该是要修正边界+1
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"EDU-ORG","错位修正")
            # 第二种情况 如果work-org紧接着edu-degree 改为edu-org
            elif nextTag.startswith("EDU") and nextTagIndex == end_index + 1:
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"EDU-ORG","错位修正")
            # 第三种情况前后都不是work改为org
            elif not lastTag.startswith("WORK-") and not nextTag.startswith("WORK-"):
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"ORG","错位修正")
            # 如果上面三格内是edu-endyear，那么work-ORG改为edu-ORG
            elif lastTag.startswith("EDU-ENDYEAR") and begin_index - lastTagIndex < 3:
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"EDU-ORG","错位修正")
            
            # 将目前的标签定位上一次标签
            lastTag = tags[i]
            # 上一次标签结束地方为end_index
            lastTagIndex = end_index
            # i定位下一个标签的起始地方
            i = nextTagIndex
        # 1.被work包含的EDU-ORG改为WORK-ORG
        # 2.紧跟着work-depart的改为work-org
        # 3.两边都不是edu的edu-org删掉
        elif tags[i] == "EDU-ORG":
            begin_index,end_index = get_begin_index_and_end_index(tags,i)
            nextTag,nextTagIndex = get_next_tag(tags,end_index + 1,tags[i])
            if debug:
                print(begin_index,end_index,nextTag,nextTagIndex)
            if lastTag.startswith("WORK-") and nextTag.startswith("WORK-"):
                # modifyTag方法传入的右边界应该是要修正边界+1
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"WORK-ORG","错位修正")
            elif nextTagIndex - end_index == 1 and nextTag.startswith("WORK-DEPART"):
                # modifyTag方法传入的右边界应该是要修正边界+1
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"WORK-ORG","错位修正") 
            elif nextTagIndex - end_index == 1 and nextTag.startswith("TITLE"):
                # modifyTag方法传入的右边界应该是要修正边界+1
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"ORG","错位修正")                       
            elif not lastTag.startswith("EDU-") and not nextTag.startswith("EDU-"):
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"ORG","错位修正")
            # 将目前的标签定位上一次标签
            lastTag = tags[i]
            # 上一次标签结束地方为end_index
            lastTagIndex = end_index
            # i定位下一个标签的起始地方
            i = nextTagIndex
        # 被edu包含的WORK-STARTTIME改为EDU-ENDYEAR，辖三个内是edu改为edu-endyear,下三格内是cre打头的改为CRE-YEAR
        elif tags[i] == "WORK-STARTTIME":
            begin_index,end_index = get_begin_index_and_end_index(tags,i)
            nextTag,nextTagIndex = get_next_tag(tags,end_index + 1,tags[i])
            if debug:
                print(begin_index,end_index,nextTag,nextTagIndex)
            # 第一种情况,上一个tag是EDU打头，下一个tag是EDU-ENDYEAR，那么tag改为EDU-STARTYEAR
            if lastTag.startswith("EDU-") and nextTag.startswith("EDU-ENDYEAR"):
                # modifyTag方法传入的右边界应该是要修正边界+1
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"EDU-STARTYEAR","错位修正")
            # 第二种情况,上一个tag是EDU-STARTYEAR,下一个是EDU，那么tag改为EDU-ENDYEAR
            elif lastTag.startswith("EDU-STARTYEAR") and nextTag.startswith("EDU-"):
                # modifyTag方法传入的右边界应该是要修正边界+1
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"EDU-ENDYEAR","错位修正")
            # 接下来三格以内是edu的改为edu-endyear
            elif nextTagIndex - end_index < 3 and nextTag.startswith("EDU-"):
                # modifyTag方法传入的右边界应该是要修正边界+1
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"EDU-ENDYEAR","错位修正")
            elif nextTagIndex - end_index < 3 and nextTag.startswith("CRE-"):
                # modifyTag方法传入的右边界应该是要修正边界+1
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"CRE-YEAR","错位修正")
            # 将目前的标签定位上一次标签
            lastTag = tags[i]
            # 上一次标签结束地方为end_index
            lastTagIndex = end_index
            # i定位下一个标签的起始地方
            i = nextTagIndex
        # 接下来是一格是EDU的ORG改为EDU-ORG 可能是edu-degree 也可能是edu-department
        elif tags[i] == "ORG":
            begin_index,end_index = get_begin_index_and_end_index(tags,i)
            nextTag,nextTagIndex = get_next_tag(tags,end_index + 1,tags[i])
            if debug:
                print(begin_index,end_index,nextTag,nextTagIndex)
            if lastTag.startswith("EDU-") and nextTag.startswith("EDU-"):
                # modifyTag方法传入的右边界应该是要修正边界+1
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"EDU-ORG","错位修正")
            elif lastTag.startswith("WORK-") and nextTag.startswith("WORK-"):
                # modifyTag方法传入的右边界应该是要修正边界+1
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"WORK-ORG","错位修正")
            elif nextTagIndex - end_index == 1 and nextTag.startswith("EDU"):
                # modifyTag方法传入的右边界应该是要修正边界+1
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"EDU-ORG","错位修正")
            
            # 将目前的标签定位上一次标签
            lastTag = tags[i]
            # 上一次标签结束地方为end_index
            lastTagIndex = end_index
            # i定位下一个标签的起始地方
            i = nextTagIndex
        # 加入规则5.1针对所有edu-degree，如果前后都没edu就清空
        # 加入规则5.2针对所有edu-degree，如果后面是导师二字就清空
        elif tags[i] == "EDU-DEGREE":
            begin_index,end_index = get_begin_index_and_end_index(tags,i)
            nextTag,nextTagIndex = get_next_tag(tags,end_index + 1,tags[i])
            if debug:
                print(begin_index,end_index,nextTag,nextTagIndex)
            if not lastTag.startswith("EDU-") and not nextTag.startswith("EDU-"):
                # modifyTag方法传入的右边界应该是要修正边界+1
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"","错位修正")
            elif introducation[end_index+1:end_index+3] == "导师":
                # modifyTag方法传入的右边界应该是要修正边界+1
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"","错位修正")       
            # 将目前的标签定位上一次标签
            lastTag = tags[i]
            # 上一次标签结束地方为end_index
            lastTagIndex = end_index
            # i定位下一个标签的起始地方
            i = nextTagIndex
        # 加入规则6.1针对所有的work-depart，如果前一位置是org则修改为depart
        elif tags[i] == "WORK-DEPART":
            begin_index,end_index = get_begin_index_and_end_index(tags,i)
            nextTag,nextTagIndex = get_next_tag(tags,end_index + 1,tags[i])
            if debug:
                print(begin_index,end_index,nextTag,nextTagIndex)
            if lastTag.startswith("ORG") and i - lastTagIndex < 3:
                # modifyTag方法传入的右边界应该是要修正边界+1
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"DEPART","错位修正")            
            # 将目前的标签定位上一次标签
            lastTag = tags[i]
            # 上一次标签结束地方为end_index
            lastTagIndex = end_index
            # i定位下一个标签的起始地方
            i = nextTagIndex
        # 加入规则7.1针对所有work-post 如果前一位置是depart则修改为post
        elif tags[i] == "WORK-POST":
            begin_index,end_index = get_begin_index_and_end_index(tags,i)
            nextTag,nextTagIndex = get_next_tag(tags,end_index + 1,tags[i])
            if debug:
                print(begin_index,end_index,nextTag,nextTagIndex)
            if lastTag.startswith("DEPART") and i - lastTagIndex < 3:
                # modifyTag方法传入的右边界应该是要修正边界+1
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"POST","错位修正")
            elif lastTag.startswith("ORG") and i -lastTagIndex < 3:                   
                # modifyTag方法传入的右边界应该是要修正边界+1
                modifyTag(debug,introducation,tags,begin_index,end_index + 1,"POST","错位修正")             
            # 将目前的标签定位上一次标签
            lastTag = tags[i]
            # 上一次标签结束地方为end_index
            lastTagIndex = end_index
            # i定位下一个标签的起始地方
            i = nextTagIndex
        else:      
            lastTag = tags[i]
            lastTagIndex = i
            i += 1
    return tags

# 返回tag的开始位置和结束位置
def get_begin_index_and_end_index(tags,begin_index):
    end_index = begin_index
    while end_index < len(tags) and tags[end_index] == tags[begin_index]:
        end_index += 1
    return begin_index,end_index-1

# 如果不存在下一个标签，返回nextTag为空，nextTagIndex为当前tag结束后的第一个位置
def get_next_tag(tags,begin_index,tag):
    nextTag = ""
    nextTagIndex = begin_index
    for i in range(begin_index,len(tags)):
        if tags[i] != tag and tags[i] != "":
            nextTag = tags[i]
            nextTagIndex = i
            break
    return nextTag,nextTagIndex

# 打印专家信息
def print_expert_info(expertInfo):
    print("基本信息如下:")
    print(expertInfo.basicInfo)
    print("工作经历如下:")
    for i in range(len(expertInfo.workInfos)):
        print(i,expertInfo.workInfos[i])
    print("教育经历如下:")
    for i in range(len(expertInfo.educationInfos)):
        print(i,expertInfo.educationInfos[i])
    print("荣誉获奖如下:")
    for i in range(len(expertInfo.creditInfos)):
        print(i,expertInfo.creditInfos[i])

# 获得中国所有省份名字典,key为简称 value为全称
def get_all_cn_province_map():
    all_cn_province_map = {}
    # 23个省    
    all_cn_province_list = ["河北省","河南省","湖北省","湖南省","台湾省",
                             "辽宁省","吉林省","陕西省","江苏省","浙江省",
                             "安徽省","福建省","江西省","山东省","山西省",
                             "广东省","海南省","四川省","贵州省","云南省",
                             "青海省","甘肃省","黑龙江省"]
    for cn_province in all_cn_province_list:
        all_cn_province_map[cn_province[:-1]] = cn_province
    
    # 4个直辖市
    all_cn_municipality_list = ["上海市","北京市","天津市","重庆市"]
    for cn_province in all_cn_municipality_list:
        all_cn_province_map[cn_province[:-1]] = cn_province
    
    # 2个特别行政区
    all_cn_special_administrative_region_list = ["香港特别行政区","澳门特别行政区"]
    for cn_province in all_cn_special_administrative_region_list:
        all_cn_province_map[cn_province] = cn_province[:-5]
    
    # 5个自治区
    all_cn_autonomous_region_list = ["宁夏回族自治区","西藏自治区","新疆维吾尔自治区","内蒙古自治区","广西壮族自治区"]
    all_cn_province_map["宁夏回族自治区"] = "宁夏"
    all_cn_province_map["西藏自治区"] = "西藏"
    all_cn_province_map["新疆维吾尔自治区"] = "新疆"
    all_cn_province_map["内蒙古自治区"] = "内蒙古"
    all_cn_province_map["广西壮族自治区"] = "广西"
    return all_cn_province_map

# 获得中国所有市级城市名字典，key为简称，value为全程
def get_all_cn_city_map():
    all_cn_city_map = {}
    cn_city_list = ["绍兴市","温州市","宁波市","杭州市","台州市","丽水市","金华市","嘉兴市","湖州市","舟山市",
                    "苏州市","南京市","常州市","无锡市","深圳市","大连市","武汉市","乐清市","衢州市","南通市",
                    "镇江市","青岛市","龙岩市","济南市","兰州市","西安市","玉林市","珠海市","兰溪市","沧州市",
                    "上虞市","合肥市","舟山市","福州市","义乌市","茂名市","富阳市","唐山市","泰安市","温岭市",
                    "诸暨市","永康市","佳木斯市","中山市","湛江市","新泰市","徐州市","盐城市","威海市","沈阳市",
                    "明光市","芜湖市","金坛市","吉林市",    
                    ]
                    #  
                    # "无棣县","仙居县","青田县","绍兴县","新昌县","文成县",
                    # "宝山区","松江区"
              
    for cn_city in cn_city_list:
        all_cn_city_map[cn_city[:-1]] = cn_city
    return all_cn_city_map

# 得到一个字典,key是所有国家中文名,value是所有国家英文名
def get_all_country_map(filename="/home/cly/poc/dict/all_country_names.txt"):
    all_country_map = {}
    f = open(filename)              
    line = f.readline()              
    while line: 
        line = line.strip()
        if line != "":
            index = line.find(" ")
            if index > 0:
                all_country_map[line[0:index]] = line[index+1:]
            else:
                all_country_map[line] = "unknown_en_name"    
        line = f.readline()  
    f.close() 
    return all_country_map



# 得到一个列表，里面是所有可能的外国地点，如果被预测为籍贯就要删除
def get_delete_native_place_list(filename="/home/cly/poc/dict/delete_native_palce_list.txt"):
    delete_native_place_list = []
    f = open(filename)              
    line = f.readline()              
    while line: 
        line = line.strip()
        delete_native_place_list.append(line)
        line = f.readline()  
    f.close() 
    return delete_native_place_list

# 得到所有学科领域列表
def get_all_major_list():
    all_major_list =["计算机科学","物理学","数学","经济学","生物学","生物化学","化学","分子生物学","计算机","社会学",
                      "遗传学","理论物理学","机械工程","免疫学","微生物学","有机化学","生态学","电气工程","应用数学","电子",
                      "统计学","心理学","人类学","天体物理学","物理","物理化学","泌尿外科","妇产科","化学工程","神经学",
                      "生物工程","生物物理学","生物医学工程","病毒学","神经科学","药理学","内科","凝聚态物理","生理学","发育生物学",
                      "进化生物学","民商法","生物","车辆工程","天文学","土木工程","地球物理学","地质学","神经生物学","生物医学",
                      "无机化学","哲学","呼吸内科","医学","地理学","法理学","管理学","粒子物理学","临床医学","细胞生物学","植物学",
                      "病理学","耳鼻喉科","分析化学","机器学习","计算机视觉","计算数学","教育学","神经内科","生物化工","心血管内科",
                      "药物化学","政治学","航天航空","机械设计","计量经济学","纳米材料","西方经济学","消化内科","园艺学","电气",
                      "电子工程系","耳鼻咽喉科","固体力学","光电子学","化学生物学","结构生物学","金融","金融学","经济法","力学",
                      "历史学","量子光学","麻醉学","大气科学","地球化学","光学工程","核医学","基础数学","精神病学","考古学",
                      "理论物理","人工智能","认知心理学","社会心理学","神经外科","生命科学","生物信息学","植物","肿瘤学","材料科学",
                      "电子学","分子遗传学","工程力学","基因组学","昆虫学","理论经济学","流行病学","内分泌","认知科学","软件工程",
                      "神经病学","数学物理学","外科学","应用化学","应用语言学","语言学","植物病理学","植物生物学","表现遗传学","产业经济学",
                      "传感器","国际法","固体物理学","环境工程","环境科学","建筑学","结构工程","金属材料","经济地理学","老年医学",
                      "粒子物理","流体力学","皮肤科","生物材料","图像处理","数理统计","刑法学","眼科","预防医学","肿瘤免疫学","水文学",
                      "电子工程","电子信息","动物学","儿科","法学","放射学","计算生物学","金融数学","概率与统计","高分子","公共事业管理",
                      "光学","海洋工程","汉语言文字学","环境生物学","会计学","民商法学","经济","经济法学","考古","计算机网络","马克思主义"
                     ]
    return all_major_list

# 得到所有职称名列表
def get_all_title_list():
    all_title_list =[ "教授级高级工程师","教授级高工","研究员级高级工程师","教授特许工程师",
                      "正高级工程师","高级工程师","中级工程师","副总工程师","总工程师","首席工程师","工程师",
                      "二级教授","副教授","特聘教授","特任教授","首席教授",
                      "终身正教授","终身教授","客座教授","助理教授","名誉教授","兼职教授","讲座教授","讲习教授","教授",
                      "青年副研究员","副研究员","助理研究员","高级研究员","首席研究员","特聘研究员","实习研究员","青年研究员","特别研究员","资深研究员","研究员",
                      "副主任医师","主任医师","主治医师","主管医师","住院医师","医师",
                      "副主任法医师","主任法医师","法医师",
                      "副主任中医师","主任中医师","中医师",
                      "副主任护师","主任护师","护师",
                      "副主任药师","主任药师","药师",
                      "高级技师","副主任技师","主任技师","技师",
                      "正高级会计师","高级会计师","助理会计师","会计师","会计员",
                      "正高级实验师","高级实验师","助理实验师","实验师","实验员",                
                      "高级讲师","助理讲师","讲师",
                      "正高级教师","高级教师","一级教师","二级教师","三级教师",
                      "经济师",
                      "国家二级心理咨询师","心理咨询师",
                      "二级鉴定官"
                    ]
    return all_title_list

def get_all_title_list_with_priorty():
    all_title_list_with_priorty = []
    title_list1  = ["正高级教师","高级教师","一级教师","二级教师","三级教师"]
    title_list2  = ["教授级高级工程师","教授级高工","教授特许工程师","总工程师","副总工程师","首席工程师","正高级工程师","研究员级高级工程师","高级工程师","中级工程师","工程师"]
    title_list3  = ["终身正教授","终身教授","教授","副教授","特聘教授","特任教授","首席教授","客座教授","助理教授","名誉教授","兼职教授","讲座教授","讲习教授"]
    title_list4  = ["高级研究员","首席研究员","副研究员","青年副研究员","助理研究员","特聘研究员","实习研究员","青年研究员","特别研究员","资深研究员","研究员"]
    title_list5  = ["主任医师","副主任医师","主治医师","主管医师","住院医师","医师"]
    title_list6  = ["主任法医师","副主任法医师","法医师"]
    title_list7  = ["主任中医师","副主任中医师","中医师"]
    title_list8  = ["主任护师","副主任护师","护师"]
    title_list9  = ["主任药师","副主任药师","药师"]
    title_list10 = ["主任技师","副主任技师","高级技师","技师"]
    title_list11 = ["正高级会计师","高级会计师","会计师","助理会计师","会计员"]
    title_list12 = ["正高级实验师","高级实验师","实验师","助理实验师","实验员"]
    title_list13 = ["高级讲师","助理讲师","讲师"]
    title_list14 = ["经济师"]
    title_list15 = ["国家二级心理咨询师","心理咨询师"]
    title_list16 = ["二级鉴定官"]
    all_title_list_with_priorty.append(title_list1)
    all_title_list_with_priorty.append(title_list2)
    all_title_list_with_priorty.append(title_list3)
    all_title_list_with_priorty.append(title_list4)
    all_title_list_with_priorty.append(title_list5)
    all_title_list_with_priorty.append(title_list6)
    all_title_list_with_priorty.append(title_list7)
    all_title_list_with_priorty.append(title_list8)
    all_title_list_with_priorty.append(title_list9)
    all_title_list_with_priorty.append(title_list10)
    all_title_list_with_priorty.append(title_list11)
    all_title_list_with_priorty.append(title_list12)
    all_title_list_with_priorty.append(title_list13)
    all_title_list_with_priorty.append(title_list14)
    all_title_list_with_priorty.append(title_list15)
    all_title_list_with_priorty.append(title_list16)
    return all_title_list_with_priorty

# 得到所有职务名列表
def get_all_post_list():
    # 上面是可能出现在荣誉奖项名称中的职务，下面是不可能出现的
    all_post_list =    [ "博士生导师","硕士生导师","研究生导师","导师","博导","硕导",
                         "辅导员",
                         "首席科学家","科学家",                               
                         "专任教师","教师",
                         "副校长","校长",
                         "学院院长","副院长","院长",
                         "教研室主任","副主任","系主任","副系主任","科主任","主任",
                         "院长助理","助理"
                         
                         "副主席","主席",
                         "副理事长","理事长","常务理事","理事",
                         "会长","副会长",
                         "副秘书长","秘书长","秘书",
                         "副所长","所长",                                         
                         "总经理","副总经理","经理",
                         "专家",
                         "委员",                 
                         "技术总监","总监",  
                         "副总裁","总裁",
                         "党委书记","党委副书记","书记",                   
                         "负责人",
                         "课题组长","组长",
                         "副处长","处长"    
                        ]
    return all_post_list

# 得到所有头衔名列表
def get_all_touxian_list():
    all_touxian_list = ["院士","会士","Fellow"]
    return all_touxian_list

# 得到所有荣誉名列表
def get_all_rongyu_list():
    all_rongyu_list =  ["科技进步","自然科学","科学技术","教学成果","杰出青年","三八红旗手","曙光学者","长江学者","浦江人才","领军人才","优秀人才"]
    return all_rongyu_list

# 得到荣誉类型字典
def get_credit_category_map(filename = "/home/cly/poc/dict/credit_category.txt"):
    credit_category_map = {}
    f = open(filename)              
    line = f.readline()              
    while line: 
        line = line.strip()
        if line != "":
            category = line.split(":")
            credit_category_map[category[0]] = category[1]           
        line = f.readline()  
    f.close() 
    return credit_category_map

# 得到荣誉级别列表
def get_credit_level_list():
    credit_level_list = ["国家级","省部级","部省级","部委级","部级","省市级","省级","市级","校级","院级","世界级","厅级","局级","县级"]
    return credit_level_list

# 得到荣誉等第列表
def get_credit_grade_list():
    credit_grade_list = ["二等奖和三等奖","一等奖和二等奖","一等奖和三等奖",
                         "二等奖","一等奖","三等奖","四等奖",
                         "特等奖","金奖","银奖","铜奖",                   
                         "一等功","二等功","三等功",
                         "第一名","第二名","第三名",
                         "提名奖","优秀奖","入围奖","最高奖","个人奖","荣誉奖","优胜奖",
                         "first prize","second prize","third prize"
                        ]  
    return credit_grade_list

# 删除荣誉奖励等于年份的奖励 1900-2050
def delete_year_credit(credit):
    if str_is_number(credit) and len(credit) == 4 and int(credit) > 1900 and int(credit) < 2050:
        return ""
    else:
        return credit

# 如果荣誉奖励以2015年 或者2020年度开头,进行拆分
def revise_credit_with_year(credit,year):
    if len(credit) >=6 and str_is_number(credit[0:4]) and credit[4:6] in ["年度","学年"]:
        return credit[6:],credit[0:4]
    elif len(credit) >=5 and str_is_number(credit[0:4]) and credit[4:5] == "年":
        return credit[5:],credit[0:4]
    elif len(credit) >=5 and str_is_number(credit[0:4]) and int(credit[0:4]) > 1900 and int(credit[0:4]) < 2050:
        return credit[4:],credit[0:4]
    else:    
        return credit,year

# 得到学校名列表
def get_school_list(filename="/home/cly/poc/dict/school.txt"):
    school_list = []
    f = open(filename)              
    line = f.readline()              
    while line: 
        line = line.strip()
        if line != "":
            school_list.append(line)
        line = f.readline()  
    f.close() 
    return school_list

# 得到组织名列表
def get_credit_org_list():
    credit_org_list = ["中国科学院","中国工程院",
                       "中国化学会","中国公路学会","中国商业联合会","中国药学会","中国纺织工业协会",
                       "国务院",
                       "发展中国家科学院",
                       "美国国家工程院","美国国家工程院","美国科学院","美国工程院","美国艺术与科学院","美国物理学会","美国科学促进会","美国癌症协会",
                       "英国皇家工程院","英国皇家化学会","英国皇家学会","英国医学科学院",
                       "法国科学院","法国建筑科学院",
                       "欧洲科学院",
                       "俄罗斯自然科学院","俄罗斯科学院","俄罗斯宇航科学院","俄罗斯工程院","俄罗斯联邦自然科学院",
                       "教育部","科技部","农业部","国务院","国防","军队",
                       "宝钢","申银万国","联合国",
                       "谷歌","AAAI","IEEE","ACM","IBM","MIT"
                       
                     ]
    return credit_org_list

# 得到所有获奖授予地的列表
def get_credit_place_list():
    # 一定要先加入全程，再加入简称
    # 加入所有省份全称
    place_list = list(get_all_cn_province_map().values())
    # 加入所有省份简称
    place_list.extend(list(get_all_cn_province_map().keys()))
    # 加入所有市全程
    place_list.extend(list(get_all_cn_city_map().values()))
    # 加入所有市简称
    place_list.extend(list(get_all_cn_city_map().keys()))
    # 加入所有国家名
    place_list.extend(list(get_all_country_map().keys()))
    # 加入其他一些特定的
    place_list.extend(["中华人民共和国","中华"])
    return place_list

# 得到所有可能的籍贯名列表
def get_native_place_list(filename="/home/cly/poc/dict/native_place_list.txt"):
    native_place_list = []
    f = open(filename)              
    line = f.readline()              
    while line: 
        line = line.strip()
        if line != "":
            native_place_list.append(line)
        line = f.readline()  
    f.close()

    # 加入所有省份全称
    native_place_list.extend(list(get_all_cn_province_map().keys()))
    # 加入所有省份简称
    native_place_list.extend(list(get_all_cn_province_map().values()))
    # 加入所有市全称
    native_place_list.extend(list(get_all_cn_city_map().keys()))
    # 加入所有市简称
    native_place_list.extend(list(get_all_cn_city_map().values()))
    return native_place_list

# 删掉字符串指定index前所有内容，找到指定index后第一个中文出现的位置，并且删除中间所有非中文字符串
def get_native_place_introducation_str(introducation,start_index):
    # 从start_index开始遍历
    for i in range(start_index,len(introducation)):
        # 如果当前字符为中文
        if char_is_chinese(introducation[i]):
            # 返回要求结果
            return introducation[i:]
    return ""


# 得到所有获奖授予地的字典
def get_credit_place_map():
    # 得到中国省份简称字典
    credit_cn_province_abbr_map = get_all_cn_province_map()
    # 得到中国省份全称字典
    credit_cn_province_entire_map = {v:v for k,v in credit_cn_province_abbr_map.items()} 
    # 得到中国各市简称字典
    credit_cn_city_abbr_map = get_all_cn_city_map()
    # 得到中国各市全程字典
    credit_cn_city_entire_map = {v:v for k,v in credit_cn_city_abbr_map.items()} 
    # 得到所有国家名
    all_country_map = get_all_country_map()
    # 构造荣誉获奖授予地的国家名字典
    credit_country_map ={k:k for k,v in all_country_map.items()} 
    
    # 将五个字典拼起来
    place_map = {k:v for d in [credit_cn_province_abbr_map,credit_cn_province_entire_map,
                               credit_cn_city_abbr_map, credit_cn_city_entire_map,
                               credit_country_map] for k,v in d.items()} 
    
    # 额外加入两个特殊的键值
    place_map["中华人民共和国"] = "中国"
    place_map["中华"] = "中国"
    return place_map

# 删除不合法的list
def delete_wrong_result_with_list(s,l):
    for element in l:
        if element == s:
            return ""
    return s 

# 如果以suffix_list中元素为后缀就删除，用于修正荣誉名，诸如荣誉名以中国，上海市，浙江省为后缀的情况
def delete_wrong_result_with_suffix_list(s,suffix_list):
    for suffix in suffix_list:
        if s.endswith(suffix):
            return ""
    return s

# 删除字符串末尾的字母，对于电话号码，邮编等字段有修正作用
def delete_str_end_alphabet(s):
    return s.rstrip(string.ascii_letters)

# 删除字符串开头的字母，对于电话号码等字段有修正作用
def delete_str_begin_alphabet(s):
    return s.lstrip(string.ascii_letters)

# 删除字符串开头的数字和符号对于教育经历的机构有修正作用
def delete_str_begin_number_and_symbol(s):
    delete_str = string.punctuation + punctuation + "0123456789"
    return s.lstrip(delete_str)

# 删除字符串结尾的数字和符号对于工作经历的机构，职务有修正作用
def delete_str_end_number_and_symbol(s):
    delete_str = string.punctuation + punctuation + "0123456789"
    return s.rstrip(delete_str)



# 如果字符被括号包围，那么删掉括号
def delete_str_begin_end_parentheses(s):
    if len(s) > 0 and s[0] == "(" and s[-1] == ")":
        return s[1:-1].strip()
    elif len(s) > 0 and s[0] == "（" and s[-1] == "）":
        return s[1:-1].strip()
    else:
        return s.strip()

# 删除字符串开头的汉字，对于电话号码等字段有修正作用
# 也可以用于删除英文机构名前的中文
def delete_str_begin_chinese(s):
    i = 0
    while i < len(s):
        if char_is_chinese(s[i]):
            i = i + 1
        else:
            break
    
    if i == len(s):
        return ""
    else:
        return s[i:]

# 删除字符串结尾的汉字，对于电话号码等字段有修正作用
# 也可以用于删除英文机构名后的中文
def delete_str_end_chinese(s):
    i = len(s) - 1
    while i > -1:
        if char_is_chinese(s[i]):
            i = i - 1
        else:
            break
    
    if i == -1:
        return ""
    else:
        return s[0:i+1]

# 判断字符串是否以指定子串开头，用于手机号，邮编校验
def check_str_prefix_with_list(s,check_list):
    for element in check_list:
        if s.startswith(element):
            return True
    return False

# 把中文括号都换成英文括号
def reviseCNBracketsToEN(s):
    s = s.replace("（","(")
    s = s.replace("）",")")
    s = s.replace("－","-")
    return s

# 把字符串中 list指定元素都替换成target 比如将 中英文逗号顿号都换成分号
def replace_str_with_list(s,list,target):
    for element in list:
        s = s.replace(element,target)
    return s


# 将机构中文名(英文名)分别拆分到中文空和英文空中
# 适用1.教育经历机构名
def split_org_to_cn_and_en(org):
    # 如果组织名中没有中文字符 则不要多想，直接返回
    if getStringChineseNumber(org) == 0 :
        return "",org

    # 在组织名为中文的情况下进行拆分
    start_index = 0
    end_index = 0 
    if "(" in org:
        start_index = org.find("(")
        end_index = org.find(")",start_index)
        if start_index > 1 and end_index - start_index > 2 and str_is_english(org[start_index + 1:end_index]):
            return org[:start_index],org[start_index + 1:end_index]

    elif "（" in org:
        start_index = org.find("（")
        end_index = org.find("）",start_index)
        if start_index > 1 and end_index - start_index > 2 and str_is_english(org[start_index + 1:end_index]):
         return org[:start_index],org[start_index + 1:end_index]


    # 最后一种情况 组织名无法拆分
    if str_is_chinese(org):
        return org,""
    else:
        return "",org
